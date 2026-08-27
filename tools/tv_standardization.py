"""TV Data Standardization — normalizes the TV team's raw spot-level files (ingest +
clean, spec step 1 of the manual process in CONSOLIDATED TV DATA.xlsx) and returns a
standardized version, per product / consolidated / both, on the user's choice.

The central transformation is converting AFFID DATE (text like "JUN28") into a real
date: the platform dates its records by the affidavit date (when the station certified
the spot aired), not by the planned spot date. Verified on the real files, aggregating
by (Product Code, Network, day): AFFID DATE gives 81.5% exact spend agreement vs 22.2%
using DATE, rising to ~95.5% once blank affidavits fall back to DATE.

Mapping tables live in tv_mappings.json so the team can edit them without touching code.

The rest of the original spec (reconciling the standardized data against the reporting
platform's export and proposing corrections) is implemented below as tested library
functions -- read_platform_export, build_reconciliation, propose_corrections,
judgment_calls, apply_corrections, build_output_workbook -- but is not wired into
render(). The tool's UI is normalize-only for now; reconciliation is a possible future
phase.
"""
import json
import os
import re
import zipfile
from datetime import date, datetime
from io import BytesIO

import openpyxl
import pandas as pd
import streamlit as st

MAPPINGS_PATH = os.path.join(os.path.dirname(__file__), "tv_mappings.json")

MONTHS_ABBR = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN",
               "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
MONTH_TO_NUM = {m: i + 1 for i, m in enumerate(MONTHS_ABBR)}

# The raw file's header row is identified by this literal in column A (spec 2.1) --
# never hardcode the row number, the report preamble's length is not guaranteed.
HEADER_MARKER = "ESTIMATE NAME"

# The 17 raw columns, in file order. NETWORK appears twice (positions 3 and 9), so
# columns are taken BY POSITION, not by name (spec 8.5).
RAW_COLUMNS = [
    "ESTIMATE NAME", "PACKAGE", "NETWORK", "QUARTER", "DAYPART", "PROGRAM NAME",
    "DATE", "MONTH", "NETWORK_SHORT", "AFFID DATE", "AFFID TIME", "LEN", "UNITS",
    "ASSIGNED GROSS", "ASSIGNED NET", "P2+ ACTIMP", "P2+ ESTIMP",
]
# Header spellings that identify the cost/impression columns, allowing for the word
# order flipping between pulls (spec 2.1: "GROSS ASSIGNED" / "NET ASSIGNED").
HEADER_ALIASES = {
    "ASSIGNED GROSS": {"ASSIGNED GROSS", "GROSS ASSIGNED"},
    "ASSIGNED NET": {"ASSIGNED NET", "NET ASSIGNED"},
    "P2+ ACTIMP": {"P2+ ACTIMP", "ACTIMP"},
    "P2+ ESTIMP": {"P2+ ESTIMP", "ESTIMP"},
}

DATA_COLUMNS = [
    "ESTIMATE NAME", "DAYPART", "PACKAGE", "NETWORK", "CLEAN NETWORK", "QUARTER",
    "DAYPART_RAW", "PROGRAM NAME", "DATE", "MONTH", "NETWORK_SHORT", "AFFID DATE",
    "AFFID TIME", "LEN", "UNITS", "ASSIGNED GROSS", "ASSIGNED NET", "P2+ ACTIMP",
    "P2+ ACTIMP*1000", "P2+ ESTIMP", "P2+ ESTIMP*1000", "AFFID MONTH",
    "PRODUCT CODE", "AUDIENCE", "EFFECTIVE DATE",
]

ACTION_ZERO_OUT = "Zero out spend"
ACTION_BACKFILL_ACTUAL = "Backfill with actual"
ACTION_BACKFILL_ESTIMATED = "Backfill with estimated"

STATE_DEFAULTS = {
    "tv_data": None,             # {product_code: normalized DataFrame}, after re-pull resolution
    "tv_file_summaries": [],
    "tv_warnings": [],
    "tv_errors": [],
    "tv_output": None,           # {"bytes", "filename", "mime"} ready to download
}


def init_state():
    for key, value in STATE_DEFAULTS.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_all():
    for key, value in STATE_DEFAULTS.items():
        st.session_state[key] = value


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def load_mappings(path: str = MAPPINGS_PATH) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def normalize_header(value) -> str:
    """Strip, upper-case and collapse whitespace before comparing column names."""
    return re.sub(r"\s+", " ", str(value if value is not None else "")).strip().upper()


# ---------------------------------------------------------------------------
# Step 1 — ingest and normalize the TV team's raw files
# ---------------------------------------------------------------------------
def product_code_from_filename(filename: str) -> str:
    """Product code is the filename's leading token -- it is not a column in the raw
    file (spec 3.2). "DHC 4.1-6.30 TV Data.xlsx" -> "DHC"."""
    stem = os.path.splitext(os.path.basename(filename))[0]
    token = re.split(r"[\s_\-.]+", stem.strip())[0]
    return token.upper()


def locate_header_row(ws) -> int:
    """1-based index of the real header row: the one whose column A is 'ESTIMATE NAME'.
    The report preamble's length is not guaranteed, so this scans rather than assuming
    row 33 (spec 2.1)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=200, max_col=1, values_only=True), start=1):
        if normalize_header(row[0]) == HEADER_MARKER:
            return i
    raise ValueError(f"Could not find the header row (column A = '{HEADER_MARKER}').")


def parse_affid_month(affid_date, month_fallback) -> str:
    """AFFID MONTH = first 3 letters of AFFID DATE, falling back to MONTH when the
    affidavit is blank (spec 3.4). Blanks arrive as ' ', not as an empty cell."""
    affid = str(affid_date or "").strip().upper()
    if len(affid) >= 3 and affid[:3] in MONTH_TO_NUM:
        return affid[:3]
    return str(month_fallback or "").strip().upper()


def effective_date(affid_date, planned_date):
    """The date the platform files a spot under (spec 3.5).

    AFFID DATE is MMMDD with no year, so the year is inherited from the planned DATE.
    If that lands more than ~45 days away from the plan, the affidavit crossed a year
    boundary and the year is nudged accordingly.
    """
    if planned_date is None or (isinstance(planned_date, float) and pd.isna(planned_date)):
        return pd.NaT
    planned = pd.to_datetime(planned_date, errors="coerce")
    if pd.isna(planned):
        return pd.NaT
    planned = planned.date() if hasattr(planned, "date") else planned

    affid = str(affid_date or "").strip().upper()
    if len(affid) < 4 or affid[:3] not in MONTH_TO_NUM:
        return pd.Timestamp(planned)

    try:
        day = int(affid[3:])
        month = MONTH_TO_NUM[affid[:3]]
        candidate = date(planned.year, month, day)
    except ValueError:
        return pd.Timestamp(planned)

    delta = (candidate - planned).days
    if delta < -45:
        candidate = date(planned.year + 1, month, day)
    elif delta > 45:
        candidate = date(planned.year - 1, month, day)
    return pd.Timestamp(candidate)


def read_tv_file(file_bytes: bytes, filename: str, mappings: dict):
    """Reads one raw TV file into the normalized DATA shape.

    Returns (DataFrame, warnings). Raises ValueError on unmapped networks or a missing
    header row -- an unmapped network would silently vanish from the join and unbalance
    the whole reconciliation (spec 7.2).
    """
    warnings = []
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        header_row = locate_header_row(ws)
        rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    finally:
        wb.close()

    if not rows:
        raise ValueError("The sheet has no rows after the header.")

    header = [normalize_header(v) for v in rows[0]]
    body = [r for r in rows[1:] if any(v is not None and str(v).strip() != "" for v in r)]
    if not body:
        raise ValueError("The sheet has a header but no data rows.")

    n_cols = len(RAW_COLUMNS)
    if len(header) < n_cols:
        raise ValueError(f"Expected at least {n_cols} columns, found {len(header)}.")

    # Columns are taken BY POSITION (NETWORK appears twice); the header is only used to
    # sanity-check that this really is the layout we expect.
    for target, position in (("ASSIGNED GROSS", 13), ("ASSIGNED NET", 14),
                             ("P2+ ACTIMP", 15), ("P2+ ESTIMP", 16)):
        found = header[position] if position < len(header) else ""
        if found not in HEADER_ALIASES[target]:
            warnings.append(
                f"{filename}: column {position + 1} is '{found}', expected one of "
                f"{sorted(HEADER_ALIASES[target])} -- read by position anyway."
            )

    df = pd.DataFrame([r[:n_cols] for r in body], columns=RAW_COLUMNS)

    out = pd.DataFrame(index=df.index)
    estimates = mappings["estimate_names"]
    networks = mappings["networks"]

    raw_estimate = df["ESTIMATE NAME"].map(lambda v: str(v).strip() if v is not None else "")
    unknown_estimates = sorted(set(raw_estimate) - set(estimates) - {""})
    if unknown_estimates:
        raise ValueError(
            f"ESTIMATE NAME values not in the mapping table: {unknown_estimates}. "
            "Add them to tv_mappings.json."
        )
    out["ESTIMATE NAME"] = raw_estimate.map(lambda v: estimates.get(v, {}).get("clean", v))
    out["DAYPART"] = raw_estimate.map(lambda v: estimates.get(v, {}).get("daypart", ""))

    out["PACKAGE"] = df["PACKAGE"]
    raw_network = df["NETWORK"].map(lambda v: str(v).strip() if v is not None else "")
    unknown_networks = sorted(set(raw_network) - set(networks) - {""})
    if unknown_networks:
        raise ValueError(
            f"NETWORK values not in the mapping table: {unknown_networks}. "
            "Add them to tv_mappings.json (an unmapped network would silently drop out "
            "of the reconciliation)."
        )
    out["NETWORK"] = raw_network
    out["CLEAN NETWORK"] = raw_network.map(networks)

    out["QUARTER"] = df["QUARTER"]
    out["DAYPART_RAW"] = df["DAYPART"]
    out["PROGRAM NAME"] = df["PROGRAM NAME"]
    out["DATE"] = pd.to_datetime(df["DATE"], format="%m/%d/%y", errors="coerce")
    out["MONTH"] = df["MONTH"].map(lambda v: str(v).strip().upper() if v is not None else "")
    out["NETWORK_SHORT"] = df["NETWORK_SHORT"]
    out["AFFID DATE"] = df["AFFID DATE"].map(lambda v: str(v).strip() if v is not None else "")
    out["AFFID TIME"] = df["AFFID TIME"].map(lambda v: str(v).strip() if v is not None else "")
    out["LEN"] = pd.to_numeric(df["LEN"], errors="coerce")
    out["UNITS"] = pd.to_numeric(df["UNITS"], errors="coerce").fillna(0)
    out["ASSIGNED GROSS"] = pd.to_numeric(df["ASSIGNED GROSS"], errors="coerce").fillna(0.0)
    out["ASSIGNED NET"] = pd.to_numeric(df["ASSIGNED NET"], errors="coerce").fillna(0.0)
    out["P2+ ACTIMP"] = pd.to_numeric(df["P2+ ACTIMP"], errors="coerce").fillna(0.0)
    out["P2+ ESTIMP"] = pd.to_numeric(df["P2+ ESTIMP"], errors="coerce").fillna(0.0)
    # Impressions arrive in thousands (spec 3.6)
    out["P2+ ACTIMP*1000"] = out["P2+ ACTIMP"] * 1000
    out["P2+ ESTIMP*1000"] = out["P2+ ESTIMP"] * 1000

    out["AFFID MONTH"] = [
        parse_affid_month(a, m) for a, m in zip(out["AFFID DATE"], out["MONTH"])
    ]
    out["PRODUCT CODE"] = product_code_from_filename(filename)
    out["AUDIENCE"] = "P2+"
    out["EFFECTIVE DATE"] = [
        effective_date(a, d) for a, d in zip(out["AFFID DATE"], out["DATE"])
    ]

    n_bad_dates = int(out["DATE"].isna().sum())
    if n_bad_dates:
        warnings.append(f"{filename}: {n_bad_dates} rows have an unparseable DATE.")

    return out[DATA_COLUMNS].reset_index(drop=True), warnings


def pull_timestamp(filename: str):
    """Best-effort 'when was this pulled' from the filename, for re-pull resolution
    (spec 3.3). Recognises a trailing M.D or M.DD token, e.g. 'VIC June Data 7.22'."""
    stem = os.path.splitext(os.path.basename(filename))[0]
    matches = re.findall(r"(?<!\d)(\d{1,2})\.(\d{1,2})(?!\d)", stem)
    if not matches:
        return None
    month, day = matches[-1]
    try:
        return (int(month), int(day))
    except ValueError:
        return None


def resolve_repulls(files: list):
    """files: [(filename, DataFrame)]. When a product code appears in more than one
    file, keep the most recent pull and report the rest (spec 3.3).

    The consolidated workbook kept the OLDER VIC pull, which left every VIC row with
    ACTIMP = 0 even though the re-pull had real audience data -- the spec calls that a
    slip, so this keeps the newer one and logs the swap.
    """
    by_product = {}
    for filename, df in files:
        product = df["PRODUCT CODE"].iloc[0] if len(df) else product_code_from_filename(filename)
        by_product.setdefault(product, []).append((filename, df))

    kept, discarded = [], []
    for product, entries in by_product.items():
        if len(entries) == 1:
            kept.append(entries[0])
            continue
        ranked = sorted(
            entries,
            key=lambda e: (pull_timestamp(e[0]) is not None, pull_timestamp(e[0]) or (0, 0)),
            reverse=True,
        )
        kept.append(ranked[0])
        for filename, df in ranked[1:]:
            discarded.append({
                "Product Code": product, "Discarded file": filename,
                "Kept instead": ranked[0][0], "Rows discarded": len(df),
            })
    return kept, discarded


# ---------------------------------------------------------------------------
# Step 2 — the platform export
# ---------------------------------------------------------------------------
def coerce_numeric(series: pd.Series) -> pd.Series:
    """Media_Cost / Impressions carry the literal string 'NULL' in some rows; left as-is
    they poison every sum (spec 4)."""
    cleaned = series.replace({"NULL": 0, "null": 0, "Null": 0})
    return pd.to_numeric(cleaned, errors="coerce").fillna(0.0)


def read_platform_export(file_bytes: bytes, sheet_name: str):
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, engine="openpyxl")
    required = ["Date", "Product Code", "Network_Name", "Daypart", "Media_Cost", "Impressions"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"The export sheet is missing required columns: {missing}")
    df = df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    df["Media_Cost"] = coerce_numeric(df["Media_Cost"])
    df["Impressions"] = coerce_numeric(df["Impressions"])
    return df


def month_key(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce").dt.strftime("%b")


# ---------------------------------------------------------------------------
# Step 3 — reconciliation
# ---------------------------------------------------------------------------
def tv_impressions(actimp: float, estimp: float) -> float:
    """Actual impressions win when present, otherwise estimated (spec 3.6)."""
    return actimp if actimp > 0 else estimp


def build_reconciliation(data: pd.DataFrame, export: pd.DataFrame, mappings: dict) -> pd.DataFrame:
    """Compares TV (truth) against the platform on Product x Month x Daypart x Network.

    Deltas follow the workbook's formulas: (TV / platform) - 1, so 0 means they agree,
    positive means the platform under-reports. Where the platform has nothing to divide
    by, the delta is left as NaN rather than a spreadsheet #DIV/0!.
    """
    daypart_map = mappings["daypart_to_platform"]

    tv = data.copy()
    tv["Month"] = tv["EFFECTIVE DATE"].dt.strftime("%b")
    tv["Daypart"] = tv["DAYPART"].map(lambda d: daypart_map.get(str(d).strip().upper()))
    tv_grouped = tv.groupby(
        ["PRODUCT CODE", "Month", "Daypart", "CLEAN NETWORK"], observed=True, dropna=False
    ).agg(
        TV_Cost=("ASSIGNED NET", "sum"),
        TV_Actual=("P2+ ACTIMP*1000", "sum"),
        TV_Estimated=("P2+ ESTIMP*1000", "sum"),
        TV_Spots=("UNITS", "sum"),
    ).reset_index().rename(columns={
        "PRODUCT CODE": "Product Code", "CLEAN NETWORK": "Network_Name"
    })

    pf = export.copy()
    pf["Month"] = month_key(pf["Date"])
    pf_grouped = pf.groupby(
        ["Product Code", "Month", "Daypart", "Network_Name"], observed=True, dropna=False
    ).agg(
        Platform_Cost=("Media_Cost", "sum"),
        Platform_Impressions=("Impressions", "sum"),
        Platform_Rows=("Media_Cost", "size"),
    ).reset_index()

    merged = pf_grouped.merge(
        tv_grouped, on=["Product Code", "Month", "Daypart", "Network_Name"], how="outer"
    )
    for col, default in (("Platform_Cost", 0.0), ("Platform_Impressions", 0.0),
                         ("Platform_Rows", 0), ("TV_Cost", 0.0), ("TV_Actual", 0.0),
                         ("TV_Estimated", 0.0), ("TV_Spots", 0)):
        merged[col] = merged[col].fillna(default)

    merged["TV_Impressions"] = [
        tv_impressions(a, e) for a, e in zip(merged["TV_Actual"], merged["TV_Estimated"])
    ]
    merged["Media Cost Delta"] = (
        merged["TV_Cost"] / merged["Platform_Cost"].replace(0, pd.NA) - 1
    )
    merged["Impression Delta"] = (
        merged["TV_Impressions"] / merged["Platform_Impressions"].replace(0, pd.NA) - 1
    )
    merged["Abs Delta"] = merged[["Media Cost Delta", "Impression Delta"]].abs().max(axis=1)
    return merged.sort_values("Abs Delta", ascending=False, na_position="first").reset_index(drop=True)


# ---------------------------------------------------------------------------
# Step 4 — propose corrections (the TRACKER)
# ---------------------------------------------------------------------------
PROPOSAL_COLUMNS = ["Apply", "Action", "Product Code", "Month", "Daypart", "Network_Name",
                    "TV Total", "Export Rows", "Value Per Row", "Platform Now", "Reason"]


def propose_corrections(reconciliation: pd.DataFrame, export: pd.DataFrame,
                         cost_tolerance: float = 0.01) -> pd.DataFrame:
    """Derives the unambiguous corrections from the reconciliation.

    Both rules require the platform to be reporting *nothing* where TV is clear about
    what happened -- that is what makes them safe to propose automatically:

    * Zero out spend -- the platform charges money for a (Product, Month, Network) that
      delivered no impressions on either side. Phantom spend on a network that never
      actually ran; the recurring case is INDEPENDENT TV NETWORK. Keyed WITHOUT daypart,
      matching the manual process.
    * Backfill impressions -- TV has impressions for a (Product, Month, Daypart, Network)
      where the platform reports zero. The TV total is spread evenly across the export's
      rows for that group.

    Measured against the manual corrections in CONSOLIDATED TV DATA.xlsx, both rules have
    **no false positives**: zero-out matches all 12 groups the analyst zeroed and proposes
    nothing they skipped; backfill matches 16 of the 20 they filled and proposes nothing
    extra. The 4 it does not propose are genuine judgment calls where the platform DOES
    report impressions but they disagree with TV (e.g. 830,000 vs 780,000) -- no threshold
    reproduces those choices, so they are deliberately left for the analyst to add by hand
    from the reconciliation table.
    """
    rows = []

    # --- Zero out spend, keyed on Product x Month x Network (no daypart) ---
    net_level = reconciliation.groupby(
        ["Product Code", "Month", "Network_Name"], observed=True, dropna=False
    ).agg(
        Platform_Cost=("Platform_Cost", "sum"),
        Platform_Impressions=("Platform_Impressions", "sum"),
        TV_Impressions=("TV_Impressions", "sum"),
    ).reset_index()

    export_month = export.copy()
    export_month["Month"] = month_key(export_month["Date"])

    for _, r in net_level.iterrows():
        if r["Platform_Cost"] <= cost_tolerance:
            continue
        if r["Platform_Impressions"] >= 1 or r["TV_Impressions"] >= 1:
            continue
        mask = (
            (export_month["Product Code"] == r["Product Code"])
            & (export_month["Month"] == r["Month"])
            & (export_month["Network_Name"] == r["Network_Name"])
        )
        rows.append({
            "Apply": True,
            "Action": ACTION_ZERO_OUT,
            "Product Code": r["Product Code"],
            "Month": r["Month"],
            "Daypart": "",
            "Network_Name": r["Network_Name"],
            "TV Total": 0.0,
            "Export Rows": int(mask.sum()),
            "Value Per Row": 0.0,
            "Platform Now": float(r["Platform_Cost"]),
            "Reason": "Platform charges spend, but neither side delivered any impressions",
        })

    # --- Backfill impressions, keyed WITH daypart ---
    for _, r in reconciliation.iterrows():
        if r["TV_Impressions"] <= 0 or r["Platform_Rows"] <= 0:
            continue
        if r["Platform_Impressions"] >= 1:
            continue
        rows.append({
            "Apply": True,
            "Action": (ACTION_BACKFILL_ACTUAL if r["TV_Actual"] > 0
                       else ACTION_BACKFILL_ESTIMATED),
            "Product Code": r["Product Code"],
            "Month": r["Month"],
            "Daypart": r["Daypart"],
            "Network_Name": r["Network_Name"],
            "TV Total": float(r["TV_Impressions"]),
            "Export Rows": int(r["Platform_Rows"]),
            "Value Per Row": float(r["TV_Impressions"]) / int(r["Platform_Rows"]),
            "Platform Now": float(r["Platform_Impressions"]),
            "Reason": "TV delivered impressions the platform reports as zero",
        })

    if not rows:
        return pd.DataFrame(columns=PROPOSAL_COLUMNS)
    proposals = pd.DataFrame(rows, columns=PROPOSAL_COLUMNS)
    # Zero-outs first: the two actions touch different columns, but fixing the order
    # makes the result reproducible (spec 8.9).
    proposals["_order"] = (proposals["Action"] != ACTION_ZERO_OUT).astype(int)
    return proposals.sort_values(
        ["_order", "Product Code", "Month", "Daypart", "Network_Name"]
    ).drop(columns="_order").reset_index(drop=True)


def judgment_calls(reconciliation: pd.DataFrame, relative_gap: float = 0.02) -> pd.DataFrame:
    """Groups where both sides report impressions but disagree by more than `relative_gap`.

    These are NOT proposed automatically -- the manual process corrected only a handful of
    them, with no rule that separates the ones it fixed from the ones it left. They are
    surfaced so the analyst can decide and add a correction by hand.
    """
    r = reconciliation
    both_report = (r["TV_Impressions"] > 0) & (r["Platform_Impressions"] >= 1) & (r["Platform_Rows"] > 0)
    gap = (r["TV_Impressions"] - r["Platform_Impressions"]).abs() / r["Platform_Impressions"]
    out = r[both_report & (gap > relative_gap)].copy()
    out["Impression Gap %"] = (gap[both_report & (gap > relative_gap)] * 100).round(1)
    columns = ["Product Code", "Month", "Daypart", "Network_Name", "Platform_Impressions",
               "TV_Impressions", "Impression Gap %", "Platform_Rows"]
    return out[columns].sort_values("Impression Gap %", ascending=False).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Step 5 — apply corrections
# ---------------------------------------------------------------------------
def apply_corrections(export: pd.DataFrame, actions: pd.DataFrame):
    """Applies the approved actions to a COPY of the export.

    Only Media_Cost and Impressions are ever written; every other column passes through
    untouched. Returns (clean export, per-action report).
    """
    clean = export.copy()
    clean["Impressions"] = clean["Impressions"].astype("float64")
    clean["Media_Cost"] = clean["Media_Cost"].astype("float64")
    month = month_key(clean["Date"])

    report = []
    for _, action in actions.iterrows():
        mask = (
            (clean["Product Code"] == action["Product Code"])
            & (month == action["Month"])
            & (clean["Network_Name"] == action["Network_Name"])
        )
        if action["Action"] != ACTION_ZERO_OUT:
            mask &= clean["Daypart"] == action["Daypart"]

        n_rows = int(mask.sum())
        entry = {
            "Action": action["Action"], "Product Code": action["Product Code"],
            "Month": action["Month"], "Daypart": action["Daypart"],
            "Network_Name": action["Network_Name"], "Rows Matched": n_rows,
        }

        if n_rows == 0:
            # Never a silent #DIV/0! -- say so explicitly (spec 6.3)
            report.append({**entry, "Rows Changed": 0, "Status": "No matching rows in the export"})
            continue

        if action["Action"] == ACTION_ZERO_OUT:
            before = clean.loc[mask, "Media_Cost"]
            changed = int((before.abs() > 1e-9).sum())
            clean.loc[mask, "Media_Cost"] = 0.0
            status = "Applied" if changed else "No change (already zero)"
            report.append({**entry, "Rows Changed": changed, "Status": status})
        else:
            value = float(action["TV Total"]) / n_rows
            before = clean.loc[mask, "Impressions"]
            changed = int((~((before - value).abs() < 1e-6)).sum())
            clean.loc[mask, "Impressions"] = value
            report.append({
                **entry, "Rows Changed": changed,
                "Status": f"Applied — {value:,.2f} per row",
            })

    return clean, pd.DataFrame(report)


# ---------------------------------------------------------------------------
# Output workbook
# ---------------------------------------------------------------------------
def build_standardized_workbook(data: pd.DataFrame) -> bytes:
    """A single-sheet workbook holding one product's (or the consolidated) normalized
    DATA table -- the actual deliverable of the simplified, normalize-only tool."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        data.to_excel(writer, sheet_name="DATA", index=False)
    return output.getvalue()


def build_output_workbook(data, reconciliation, tracker, clean, verification, log) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, frame in (
            ("DATA", data), ("RECONCILIATION", reconciliation), ("TRACKER", tracker),
            ("CLEAN", clean), ("VERIFICATION", verification), ("LOG", log),
        ):
            frame = frame if frame is not None and len(frame) else pd.DataFrame({"(empty)": []})
            frame.to_excel(writer, sheet_name=name[:31], index=False)
    return output.getvalue()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
def render():
    if st.button("← Back to menu"):
        st.session_state.active_tool = None
        st.rerun()

    init_state()
    mappings = load_mappings()

    st.title("TV Data Standardization")
    st.caption(
        "Upload the TV team's raw spot-level files and get back a standardized version: "
        "AFFID DATE converted to a real date, networks and dayparts mapped to the "
        "platform's naming, impressions scaled to real units."
    )

    tv_files = st.file_uploader(
        "TV team files (.xlsx) — one per product, e.g. DHC / TRE / VIC",
        type=["xlsx"], accept_multiple_files=True, key="tv_files_uploader",
    )

    if tv_files and st.button("Standardize", type="primary"):
        parsed, warnings, errors = [], [], []
        for uploaded in tv_files:
            try:
                df, file_warnings = read_tv_file(uploaded.getvalue(), uploaded.name, mappings)
                parsed.append((uploaded.name, df))
                warnings.extend(file_warnings)
            except Exception as exc:
                errors.append(f"{uploaded.name}: {exc}")

        if not parsed:
            for e in errors:
                st.error(e)
            st.error("None of the TV files could be read.")
            st.stop()

        kept, discarded = resolve_repulls(parsed)
        for d in discarded:
            warnings.append(
                f"{d['Product Code']}: using the newer pull '{d['Kept instead']}' and "
                f"discarding '{d['Discarded file']}' ({d['Rows discarded']} rows)."
            )

        data_by_product = {df["PRODUCT CODE"].iloc[0]: df for _, df in kept}
        summaries = [
            {"File": name, "Product Code": df["PRODUCT CODE"].iloc[0], "Rows": len(df),
             "Blank AFFID DATE": int((df["AFFID DATE"] == "").sum())}
            for name, df in kept
        ]

        st.session_state.tv_data = data_by_product
        st.session_state.tv_file_summaries = summaries
        st.session_state.tv_warnings = warnings
        st.session_state.tv_errors = errors
        st.session_state.tv_output = None
        st.rerun()

    data_by_product = st.session_state.tv_data
    if not data_by_product:
        return

    if st.button("Start over"):
        reset_all()
        st.rerun()

    st.divider()
    st.subheader("Files read")
    st.dataframe(pd.DataFrame(st.session_state.tv_file_summaries),
                 use_container_width=True, hide_index=True)

    total_rows = sum(len(df) for df in data_by_product.values())
    total_blank = sum(int((df["AFFID DATE"] == "").sum()) for df in data_by_product.values())
    c1, c2, c3 = st.columns(3)
    c1.metric("Products", len(data_by_product))
    c2.metric("Total spot rows", f"{total_rows:,}")
    c3.metric("Blank affidavits", f"{total_blank:,}")

    for w in st.session_state.tv_warnings:
        st.warning(w)
    for e in st.session_state.tv_errors:
        st.error(e)

    st.divider()
    st.subheader("Output format")
    output_choice = st.radio(
        "How should the standardized files be delivered?",
        ["Separate files per product", "One consolidated file", "Both"],
        key="tv_output_format",
    )

    if st.button("Generate", type="primary", use_container_width=True):
        stamp = datetime.now().strftime("%Y%m%d_%H%M")
        products = sorted(data_by_product)

        if output_choice == "One consolidated file":
            consolidated = pd.concat([data_by_product[p] for p in products], ignore_index=True)
            file_bytes = build_standardized_workbook(consolidated)
            filename = f"TV_Standardized_Consolidated_{stamp}.xlsx"
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            buf = BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
                if output_choice in ("Separate files per product", "Both"):
                    for p in products:
                        zf.writestr(
                            f"TV_Standardized_{p}_{stamp}.xlsx",
                            build_standardized_workbook(data_by_product[p]),
                        )
                if output_choice == "Both":
                    consolidated = pd.concat(
                        [data_by_product[p] for p in products], ignore_index=True
                    )
                    zf.writestr(
                        f"TV_Standardized_Consolidated_{stamp}.xlsx",
                        build_standardized_workbook(consolidated),
                    )
            file_bytes = buf.getvalue()
            filename = f"TV_Standardized_{stamp}.zip"
            mime = "application/zip"

        st.session_state.tv_output = {"bytes": file_bytes, "filename": filename, "mime": mime}
        st.rerun()

    output = st.session_state.tv_output
    if not output:
        return

    st.divider()
    st.download_button(
        f"Download {output['filename']}",
        data=output["bytes"], file_name=output["filename"], mime=output["mime"],
        type="primary", use_container_width=True,
    )
