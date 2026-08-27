"""Merit Inspect — monthly marketing data QA: row-level rule checks, delivery
(spend-vs-metric) analysis, channel/month summary and a data-coverage checklist.

Ported from the standalone script `Merit_Inspect.py` (file-picker + local `output/`
folder, run by hand) into an in-app, upload/download flow. All file I/O happens in
memory (BytesIO); nothing is written to disk. Every business rule (Knorr Product_Line,
TV audience, Twitter/X partner, placeholder placements, delivery checks, etc.) is
unchanged from the original script — only the input/output boundary changed.
"""
import os
import re
import time
from datetime import datetime
from io import BytesIO

import numpy as np
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment

# =============================================================================
# 1) FILE-LEVEL CONFIG (rarely changes — same as the standalone script)
# =============================================================================
OUTPUT_PREFIX = "QA Results"

# --- Knorr Product_Line rules (Inspect) --------------------------------------
KNORR_BRAND            = "Knorr"
KNORR_SOCIAL_CHANNEL   = "Digital Social"
KNORR_ALLOWED_PL       = ["Bouillon", "Sides"]
KNORR_POPPY_KW         = "poppy"
KNORR_BOUILLON_KW      = ["bouillon", "flavorbase", "soup", "flavor-base", "poppy"]
KNORR_SIDES_KW         = ["sides", "cups"]
KNORR_PLACEMENT_SIDES  = "mls"
KNORR_PLACEMENT_BOU    = "bou"

KNORR_TV_CHANNEL       = "TV"
KNORR_TV_DAYPART_VALUE = "Hispanic"
KNORR_COL_DAYPART      = "Daypart"

TV_METRIC_BY_BRAND = {
    "Knorr":      "Impressions",
    "Hellmann's": "GRPs",
    "Seventh Generation": "Impressions",
}
TV_METRIC_DEFAULT = None

PLACEHOLDER_SUBSTRINGS = ["Dummy", "DELETE", "DO NOT USE"]
PLACEHOLDER_ISOLATED   = ["DNU"]
FLAG_NON_4DIGIT_AUDIENCE = True

# =============================================================================
# 2) STRUCTURAL CONFIGURATION (rarely changes)
# =============================================================================
CFG = {
    "col_channel":      "Channel",
    "col_placement":    "Package_Placement_Name",
    "col_cost":         "Media_Cost",
    "col_impressions":  "Impressions",
    "col_grps":         "GRPs",
    "col_brand":        "Brand",
    "col_raw_partner":  "Raw_Partner",
    "col_daypart":      "Daypart",
    "col_network":      "Network_Name",
    "col_date":         "Date",
    "col_audience":     "Audience",
    "col_breakout":     "Breakout",
    "col_category":     "Category",
    "col_product_line": "Product_Line",
    "col_prisma":       "Prisma_Campaign_Secondary",
    "col_team":         "Team",

    "critical_always":   ["Channel", "Brand", "Product_Line", "Category",
                          "Raw_Partner", "Audience", "Retailer", "Breakout"],
    "critical_except_tv":["Campaign", "Prisma_Campaign_Secondary", "Package_Placement_Name"],
    "tv_channel_value":     "TV",
    "social_channel_value": "Digital Social",

    "grp_offline_channels":  ["Print"],
    "impr_offline_channels": ["OOH", "DOOH", "Cinema"],
    "expected_offline_channels": ["Print", "OOH", "DOOH", "Cinema"],
}
KNORR_OTHER_SAY_KW   = ["bhamp", "ihamp"]
KNORR_BRAND_SAY_KW   = ["bsay"]
KNORR_OTHER_SAY_VAL  = "Other Say"
KNORR_BRAND_SAY_VAL  = "Brand Say"

SOCIAL_CODES = {
    "FBCR","FBIF","FBSTRM","FBMW","FBR","FBTAS","FBST","FIAR","FICR","FICO","FIDPA",
    "FIIMG","FIIF","FBIE","FIMM","FIPO","FIPA","FIR","FIRO","FISO","FIVID","INFB",
    "IGCR","IGIF","IGPO","IGR","IGST","IGSR","IGTH","LICA","LIDA","LIIM","LIVI",
    "PICA","PICO","PEDP","PIIDEA","PIIP","PIMW","PIPP","PIMG","PVID","PIQA","PISP",
    "PISA","PISL","RECA","RCT","RECO","REPR","REFV","REGIF","REIF","REIC","REMT",
    "REST","RVIDE","SNAD","SNBU","SNPC","SNSPC","SNAS","SNCL","SNCO","SNCOS","SNCSG",
    "SNFI","SNLE","SNIG","SNVD","SSAR","SNSL","SNST","SOCB","TKBB","TKCA","TKCO",
    "TKDO","TKDI","TKFD","TKHC","TTIM","TKVID","TTPP","TKPS","TTPCU","TTPST","TKSE",
    "TKSP","TKTF","TKTV","TKVD","TTCSH","TWAM","TWCR","TWCD","TWFV","TWLR","TWPR",
    "TWTR","TWTS","TWPT","TWSPA","TWSPR","TWSCA","TWWC",
}
TWITTER_CODES = {
    "TWAM", "TWCR", "TWCD", "TWFV", "TWLR", "TWPR", "TWTR",
    "TWTS", "TWPT", "TWSPA", "TWSPR", "TWSCA", "TWWC",
}
X_CORP_SUFFIX          = " - X CORP"
X_CORP_RULE_SEVERITY   = "ERROR"
STD_BLOCK_REGEX = (
    r"(?:^|_)(CPC|CPM|CPV)_"
    r"(AUC|AVBC|PRGN|GURA|FEE)_"
    r"([^_]+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)_([^_]+)"
)

SEV_RED, SEV_YELLOW = "red", "yellow"
FILL_RED    = PatternFill("solid", fgColor="E25B45")  # flame pea
FILL_YELLOW = PatternFill("solid", fgColor="FAC172")  # rojah
FILL_HEADER = PatternFill("solid", fgColor="83D0C6")  # monte carlo
FILL_GREEN  = PatternFill("solid", fgColor="ADC865")  # wild willow
VERDICT_TITLE    = "Verdict"
TOP_ISSUES_TITLE = "Top issues (what to fix first)"

METRIC_CROSS_SEVERITY = "REVIEW"
CROSS_METRIC_INCLUDE_TV = True

BREAKOUT_OTHER_SAY_VALUE = "Other Say"
BREAKOUT_RULE_SEVERITY   = "ERROR"

ONLINE_PAIR_REGEX = r"(?:^|_)(Online[^_]*)_([^_]*)"
CHANNEL_TYPE_AUDIO_MARK = "Audio"
CHANNEL_TYPE_VIDEO_MARK = "Video"
CHANNEL_TYPE_CONFLICT_SEVERITY = "ERROR"

PARTNER_COMPANION_RULES = [
    {"channel": "Digital FEP", "partner": "THE TRADE DESK INC", "flag": "Missing TTD Partner"},
]
PARTNER_COMPANION_SEVERITY = "REVIEW"

TV_AUDIENCE_BY_BRAND = {
    "Hellmann's": "A2564",
    "Knorr":      "P2+",
}
TV_AUDIENCE_RULE_SEVERITY = "ERROR"
MONTH_NAMES = {1:"January",2:"February",3:"March",4:"April",5:"May",6:"June",
               7:"July",8:"August",9:"September",10:"October",11:"November",12:"December"}

DELIVERY_COLS = ["Severity", "Issue", "Channel", "Brand", "Team", "Raw_Partner",
                 "Unit_Type", "Unit", "Month",
                 "Total Impressions", "Total GRPs", "Total Cost"]

CHANNEL_TITLE = "Channel summary (by month)"
DIGITAL_CHANNELS = {
    "Digital Video", "Digital FEP", "Digital Social",
    "Digital Display", "Digital Audio",
}
MASTER_CHANNEL_DIGITAL_LABEL = "Digital"

CPM_HIGH_THRESHOLD    = 50.0
CPM_CHECK_UNIT_TYPES  = ["Placement", "TV", "Offline (Raw_Partner)"]
CPM_HIGH_SEVERITY     = "REVIEW"

PRIORITY_TITLE = "What to fix first"
_DELIVERY_CATS = [
    ("Spend without delivery",  ["spend but 0 grps", "spend but 0 impressions"], "ERROR"),
    ("Delivery without spend",  ["but 0 spend"], "REVIEW"),
    ("Metric cross-contamination", ["has impressions", "has grps"], "REVIEW"),
    ("Unmapped TV metric",      ["no defined delivery metric"], "REVIEW"),
    ("High CPM",                ["high cpm"], "REVIEW"),
    ("High CPP",                ["high cpp"], "REVIEW"),
]

RULE_LABELS = {
    "rule_negative_cost":           "Negative cost",
    "rule_social_channel_mismatch": "Social channel mismatch",
    "rule_unexpected_structure":    "Unexpected placement structure",
    "rule_placeholder_placement":   "Placeholder placement (Dummy/DELETE/DNU)",
    "rule_audience_mismatch":       "Audience mismatch (code vs column)",
    "rule_audience_code_issue":     "Audience code issue (not in catalog/format)",
    "rule_breakout_other_say":      "'Other Say' outside Digital Social",
    "rule_partner_companion":       "Partner missing companion",
    "rule_tv_audience":             "TV audience incorrect (by brand)",
    "rule_channel_type_conflict":   "Channel/format conflict (Audio/Video)",
    "rule_twitter_x_corp":          "X Corp partner missing",
    "rule_knorr_pl_error":  "Knorr Product_Line incorrect",
    "rule_knorr_pl_review": "Knorr Product_Line needs review",
    "rule_knorr_breakout_error":  "Knorr/Social Breakout incorrect",
    "rule_knorr_breakout_review": "Knorr/Social Breakout needs review",
}

STATE_DEFAULTS = {
    "inspect_results": None,   # dict produced by the "Run QA checks" step
}


def init_state():
    for key, value in STATE_DEFAULTS.items():
        if key not in st.session_state:
            st.session_state[key] = value
    if "inspect_uploader_key" not in st.session_state:
        st.session_state["inspect_uploader_key"] = 0


def reset_all():
    for key, value in STATE_DEFAULTS.items():
        st.session_state[key] = value
    # bump the uploader widgets' key so they visually clear too, not just the results
    st.session_state["inspect_uploader_key"] = st.session_state.get("inspect_uploader_key", 0) + 1


# ----------------------------------------------------------------------------
# Small helpers (unchanged logic)
# ----------------------------------------------------------------------------
def _join_distinct(s):
    vals = sorted({str(x).strip() for x in s.dropna() if str(x).strip() != ""})
    return ", ".join(vals)


def _period_to_label(p):
    if p is pd.NaT or pd.isna(p):
        return "Unknown"
    return f"{MONTH_NAMES[p.month]} {p.year}"


def _parse_dates(s: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(s):
        return s
    out = pd.to_datetime(s, format="%m/%d/%Y", errors="coerce")
    bad = out.isna() & s.notna()
    if bad.any():
        out = out.fillna(pd.to_datetime(s, errors="coerce"))
    return out


def _sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"<>|]', "_", name).strip()


def _norm_aud(x):
    if pd.isna(x):
        return ""
    s = re.sub(r"\s*:\s*", ": ", str(x))
    s = re.sub(r"\s+", " ", s)
    return s.strip().casefold()


def load_audience_catalog(uploaded_bytes: bytes = None):
    """{code(4 digits): audience name}. Uses the uploaded override if given,
    otherwise the built-in tools/config/audience_codes.csv. Returns None if
    neither is available (the audience rules then do nothing)."""
    if uploaded_bytes is not None:
        cat = pd.read_csv(BytesIO(uploaded_bytes), dtype=str).fillna("")
    else:
        default_path = os.path.join(os.path.dirname(__file__), "config", "audience_codes.csv")
        if not os.path.exists(default_path):
            return None
        cat = pd.read_csv(default_path, dtype=str).fillna("")
    cat["Code"] = cat["Code"].astype(str).fillna("").str.strip().str.zfill(4)
    cat["Audience"] = cat["Audience"].astype(str).fillna("").str.strip()
    cat = cat[cat["Code"] != ""]
    return dict(zip(cat["Code"], cat["Audience"]))


# =============================================================================
# CONTEXT: heavy computations done ONCE and shared across rules
# =============================================================================
def build_context(df: pd.DataFrame, catalog) -> dict:
    ctx = {}
    ctx["cost_num"] = pd.to_numeric(df[CFG["col_cost"]], errors="coerce")
    ctx["impr_num"] = pd.to_numeric(df[CFG["col_impressions"]], errors="coerce")
    ctx["grps_num"] = pd.to_numeric(df[CFG["col_grps"]], errors="coerce")

    dates = _parse_dates(df[CFG["col_date"]])
    period = dates.dt.to_period("M")
    ctx["dates"] = dates
    ctx["month_period"] = period
    ctx["month_label"] = period.map(_period_to_label)
    ctx["month_ts"] = period.dt.to_timestamp()

    pn = df[CFG["col_placement"]].astype("string")
    ext = pn.str.extract(STD_BLOCK_REGEX)
    ctx["parseable"] = df[CFG["col_placement"]].notna() & ext[0].notna()
    ctx["audience_code"] = ext[2].str.strip()
    inv = ext[9].str.strip().str.upper()
    ctx["inventory"] = inv
    ctx["is_social"] = inv.isin(SOCIAL_CODES)

    ctx["audience_actual"] = df[CFG["col_audience"]].astype("string")
    ctx["has_catalog"] = catalog is not None
    if catalog:
        ctx["audience_catalog_name"] = ctx["audience_code"].map(catalog).astype("string")
    else:
        ctx["audience_catalog_name"] = pd.Series(pd.NA, index=df.index, dtype="string")

    ctx["knorr_pl"] = _knorr_expected_pl(df, ctx)
    ctx["knorr_breakout"] = _knorr_expected_breakout(df, ctx)
    return ctx


# =============================================================================
# BUSINESS RULES ZONE (ROW-LEVEL) — unchanged from the standalone script
# CONTRACT: function(df, ctx) -> (mask, severity, messages)
# =============================================================================
def _make_null_rule(column: str, except_tv: bool):
    def rule(df, ctx):
        col = df[column]
        mask = col.isna() | (col.astype(str).fillna("").str.strip() == "")
        if except_tv:
            mask = mask & (df[CFG["col_channel"]] != CFG["tv_channel_value"])
        messages = pd.Series(f"Null in critical field: {column}", index=df.index)
        return mask, SEV_RED, messages
    rule.__name__ = f"null__{column}"
    return rule


def rule_negative_cost(df, ctx):
    cost = ctx["cost_num"]
    mask = cost < 0
    messages = "Negative cost: " + cost.round(2).astype(str).fillna("")
    return mask, SEV_RED, messages


def rule_social_channel_mismatch(df, ctx):
    inv, is_social, parseable = ctx["inventory"], ctx["is_social"], ctx["parseable"]
    ch = df[CFG["col_channel"]]
    is_social_channel = ch == CFG["social_channel_value"]
    mask_a = parseable & is_social & (~is_social_channel)
    mask_b = parseable & (~is_social) & is_social_channel
    mask = mask_a | mask_b
    msg_a = ("Channel mismatch: placement indicates Social (inventory '"
             + inv.astype(str).fillna("") + "') but Channel='" + ch.astype(str).fillna("") + "'")
    msg_b = ("Channel mismatch: Channel='" + CFG["social_channel_value"]
             + "' but inventory '" + inv.astype(str).fillna("") + "' is not Social")
    messages = pd.Series(np.where(mask_a, msg_a, msg_b), index=df.index)
    return mask, SEV_YELLOW, messages


def rule_unexpected_structure(df, ctx):
    pn = df[CFG["col_placement"]]
    has_value = pn.notna() & (pn.astype(str).fillna("").str.strip() != "")
    mask = has_value & (~ctx["parseable"])
    messages = pd.Series(
        "Placement Name with unexpected structure: standardized block "
        "(methodology+buy type+audience...) not found; channel/audience could not be derived",
        index=df.index)
    return mask, SEV_YELLOW, messages


def rule_placeholder_placement(df, ctx):
    up = df[CFG["col_placement"]].astype(str).fillna("").str.upper()
    found = pd.Series("", index=df.index)
    any_mask = pd.Series(False, index=df.index)
    specs = ([(t, False) for t in PLACEHOLDER_SUBSTRINGS]
             + [(t, True) for t in PLACEHOLDER_ISOLATED])
    for token, isolated in specs:
        T = token.upper()
        if isolated:
            pat = r'(?<![A-Z])' + re.escape(T) + r'(?![A-Z])'
        else:
            pat = re.escape(T)
        m = up.str.contains(pat, regex=True, na=False)
        prev = found
        new_val = pd.Series(np.where(prev == "", token, prev.astype(str).fillna("") + ", " + token),
                            index=df.index)
        found = found.mask(m, new_val)
        any_mask = any_mask | m
    messages = ("Possible placeholder placement ('" + found
                + "' found); review whether the row should be removed")
    return any_mask, SEV_YELLOW, messages


def rule_audience_mismatch(df, ctx):
    if not ctx["has_catalog"]:
        empty = pd.Series(False, index=df.index)
        return empty, SEV_RED, pd.Series("", index=df.index)
    code = ctx["audience_code"]
    is_4d = code.str.fullmatch(r"\d{4}").fillna(False)
    cat_name = ctx["audience_catalog_name"]
    found = is_4d & cat_name.notna()
    norm_cat = cat_name.map(_norm_aud)
    norm_act = ctx["audience_actual"].map(_norm_aud)
    mask = found & (norm_cat != norm_act)
    messages = ("Audience mismatch: placement code " + code.astype(str).fillna("")
                + " = '" + cat_name.astype(str).fillna("") + "' but Audience column = '"
                + ctx["audience_actual"].astype(str).fillna("") + "'")
    return mask, SEV_RED, messages


def rule_audience_code_issue(df, ctx):
    if not ctx["has_catalog"]:
        empty = pd.Series(False, index=df.index)
        return empty, SEV_YELLOW, pd.Series("", index=df.index)
    code = ctx["audience_code"]
    parseable = ctx["parseable"]
    is_4d = code.str.fullmatch(r"\d{4}").fillna(False)
    cat_name = ctx["audience_catalog_name"]
    not_found = parseable & is_4d & cat_name.isna()
    if FLAG_NON_4DIGIT_AUDIENCE:
        invalid_fmt = parseable & code.notna() & (~is_4d)
    else:
        invalid_fmt = pd.Series(False, index=df.index)
    mask = not_found | invalid_fmt
    msg_nf = ("Audience code '" + code.astype(str).fillna("")
              + "' not found in catalog (config/audience_codes.csv)")
    msg_inv = "Audience segment is not a 4-digit code: '" + code.astype(str).fillna("") + "'"
    messages = pd.Series(np.where(not_found, msg_nf, msg_inv), index=df.index)
    return mask, SEV_YELLOW, messages


def rule_breakout_other_say(df, ctx):
    breakout = df[CFG["col_breakout"]].astype(str).fillna("").str.strip().str.casefold()
    channel = df[CFG["col_channel"]]
    is_other_say = breakout == BREAKOUT_OTHER_SAY_VALUE.casefold()
    is_social = channel == CFG["social_channel_value"]
    mask = is_other_say & (~is_social)
    sev = SEV_RED if BREAKOUT_RULE_SEVERITY == "ERROR" else SEV_YELLOW
    messages = ("Breakout = '" + BREAKOUT_OTHER_SAY_VALUE + "' is only allowed for '"
                + CFG["social_channel_value"] + "', but Channel = '"
                + channel.astype(str).fillna("") + "'")
    return mask, sev, messages


def rule_partner_companion(df, ctx):
    channel = df[CFG["col_channel"]]
    partner_cf = df[CFG["col_raw_partner"]].astype(str).fillna("").str.strip().str.casefold()
    mask = pd.Series(False, index=df.index)
    msg = pd.Series("", index=df.index)
    for r in PARTNER_COMPANION_RULES:
        this = (channel == r["channel"]) & (partner_cf == r["partner"].casefold())
        text = (f"{r['flag']}: Channel '{r['channel']}' with partner "
                f"'{r['partner']}' is missing its companion "
                f"(expected e.g. '{r['partner']} - <name>')")
        mask = mask | this
        msg = msg.mask(this, text)
    sev = SEV_RED if PARTNER_COMPANION_SEVERITY == "ERROR" else SEV_YELLOW
    return mask, sev, msg


def rule_tv_audience(df, ctx):
    channel = df[CFG["col_channel"]]
    is_tv = channel == CFG["tv_channel_value"]
    brand = df[CFG["col_brand"]].astype(str).fillna("").str.strip()
    audience = df[CFG["col_audience"]].astype(str).fillna("").str.strip()
    expected = brand.map(TV_AUDIENCE_BY_BRAND)
    has_rule = is_tv & expected.notna()
    mask = has_rule & (audience.str.casefold() != expected.astype(str).fillna("").str.casefold())
    messages = ("TV audience for brand '" + brand + "' should be '"
                + expected.astype(str).fillna("") + "' but Audience = '" + audience + "'")
    sev = SEV_RED if TV_AUDIENCE_RULE_SEVERITY == "ERROR" else SEV_YELLOW
    return mask, sev, messages


def rule_channel_type_conflict(df, ctx):
    pn = df[CFG["col_placement"]].astype("string")
    ext = pn.str.extract(ONLINE_PAIR_REGEX)
    seg_a = ext[0].fillna("")
    seg_b = ext[1].fillna("")
    a_aud = seg_a.str.contains(CHANNEL_TYPE_AUDIO_MARK, case=False, na=False)
    a_vid = seg_a.str.contains(CHANNEL_TYPE_VIDEO_MARK, case=False, na=False)
    b_aud = seg_b.str.contains(CHANNEL_TYPE_AUDIO_MARK, case=False, na=False)
    b_vid = seg_b.str.contains(CHANNEL_TYPE_VIDEO_MARK, case=False, na=False)
    mask = (a_aud & b_vid) | (a_vid & b_aud)
    messages = ("Channel/format conflict in placement: '" + seg_a.astype(str).fillna("")
                + "' + '" + seg_b.astype(str).fillna("") + "' mixes Audio and Video")
    sev = SEV_RED if CHANNEL_TYPE_CONFLICT_SEVERITY == "ERROR" else SEV_YELLOW
    return mask, sev, messages


def rule_twitter_x_corp(df, ctx):
    inv = ctx["inventory"]
    is_twitter = ctx["parseable"] & inv.isin(TWITTER_CODES)
    partner = df[CFG["col_raw_partner"]].astype("string").fillna("").str.strip()
    suffix_cf = X_CORP_SUFFIX.strip().casefold()
    ends_ok = partner.str.casefold().str.strip().str.endswith(suffix_cf)
    mask = is_twitter & (~ends_ok)
    messages = ("Twitter/X inventory code '" + inv.astype(str).fillna("")
                + "' but Raw_Partner does not end with '" + X_CORP_SUFFIX
                + "' (current: '" + partner + "')")
    sev = SEV_RED if X_CORP_RULE_SEVERITY == "ERROR" else SEV_YELLOW
    return mask, sev, messages


def _knorr_expected_pl(df, ctx):
    n = len(df)
    brand = df[CFG["col_brand"]].astype("string").fillna("").str.strip()
    channel = df[CFG["col_channel"]].astype("string").fillna("").str.strip()

    is_knorr  = brand.str.casefold() == KNORR_BRAND.casefold()
    is_social = channel.str.casefold() == KNORR_SOCIAL_CHANNEL.casefold()
    is_tv     = channel.str.casefold() == KNORR_TV_CHANNEL.casefold()

    prisma    = df[CFG["col_prisma"]].astype("string").fillna("")
    creative  = df["Creative Name"].astype("string").fillna("") \
                if "Creative Name" in df.columns else pd.Series([""]*n, index=df.index)
    placement = df[CFG["col_placement"]].astype("string").fillna("")
    daypart   = df[CFG["col_daypart"]].astype("string").fillna("").str.strip()

    pr_cf = prisma.str.casefold()
    cr_cf = creative.str.casefold()
    pn_cf = placement.str.casefold()

    expected = pd.Series(pd.NA, index=df.index, dtype="string")
    status   = pd.Series("", index=df.index, dtype="string")

    social = is_knorr & is_social
    poppy = social & pr_cf.str.contains(KNORR_POPPY_KW, na=False)
    expected = expected.mask(poppy, "Bouillon"); status = status.mask(poppy, "target")

    non_poppy = social & (~pr_cf.str.contains(KNORR_POPPY_KW, na=False))
    has_cr = creative.str.strip() != ""
    cr_bou = cr_cf.apply(lambda s: any(k in s for k in KNORR_BOUILLON_KW))
    cr_sid = cr_cf.apply(lambda s: any(k in s for k in KNORR_SIDES_KW))

    m = non_poppy & (~has_cr)
    status = status.mask(m, "review:no_creative")
    m = non_poppy & has_cr & cr_bou & cr_sid
    status = status.mask(m, "review:ambiguous_creative")
    m = non_poppy & has_cr & cr_bou & (~cr_sid)
    expected = expected.mask(m, "Bouillon"); status = status.mask(m, "target")
    m = non_poppy & has_cr & cr_sid & (~cr_bou)
    expected = expected.mask(m, "Sides"); status = status.mask(m, "target")
    m = non_poppy & has_cr & (~cr_bou) & (~cr_sid)
    status = status.mask(m, "review:no_keyword")

    other = is_knorr & (~is_social) & (~is_tv)
    pn_mls = pn_cf.str.contains(KNORR_PLACEMENT_SIDES, na=False)
    pn_bou = pn_cf.str.contains(KNORR_PLACEMENT_BOU, na=False)

    m = other & pn_mls & pn_bou
    status = status.mask(m, "review:ambiguous_placement")
    m = other & pn_mls & (~pn_bou)
    expected = expected.mask(m, "Sides"); status = status.mask(m, "target")
    m = other & pn_bou & (~pn_mls)
    expected = expected.mask(m, "Bouillon"); status = status.mask(m, "target")
    m = other & (~pn_mls) & (~pn_bou)
    status = status.mask(m, "review:no_placement_kw")

    tv_hispanic = is_knorr & is_tv & \
                  (daypart.str.casefold() == KNORR_TV_DAYPART_VALUE.casefold())
    expected = expected.mask(tv_hispanic, "Bouillon")
    status = status.mask(tv_hispanic, "target")

    return pd.DataFrame({"is_knorr": is_knorr, "expected": expected,
                         "status": status}, index=df.index)


def rule_knorr_pl_error(df, ctx):
    info = ctx["knorr_pl"]
    actual = df[CFG["col_product_line"]].astype("string").fillna("").str.strip()
    allowed_cf = [v.casefold() for v in KNORR_ALLOWED_PL]

    has_target = info["status"] == "target"
    mismatch = has_target & (actual.str.casefold() != info["expected"].str.casefold())

    not_allowed = info["is_knorr"] & (~actual.str.casefold().isin(allowed_cf))

    mask = (mismatch | not_allowed).fillna(False)

    msg_mismatch = ("Knorr Product_Line should be '" + info["expected"].astype(str).fillna("")
                    + "' but is '" + actual + "'")
    msg_notallowed = ("Knorr Product_Line must be Bouillon or Sides, but is '"
                      + actual + "'")
    messages = pd.Series(np.where(mismatch, msg_mismatch, msg_notallowed),
                         index=df.index)
    return mask, SEV_RED, messages


def rule_knorr_pl_review(df, ctx):
    info = ctx["knorr_pl"]
    is_review = info["status"].str.startswith("review:").fillna(False)
    reasons = {
        "review:no_creative":          "Knorr/Social, not Poppy, and no Creative Name — can't classify",
        "review:ambiguous_creative":   "Knorr/Social: Creative Name has BOTH Bouillon and Sides keywords",
        "review:no_keyword":           "Knorr/Social: Creative Name has no known keyword",
        "review:ambiguous_placement":  "Knorr/non-Social: Placement has BOTH MLS and BOU",
        "review:no_placement_kw":      "Knorr/non-Social: Placement has neither MLS nor BOU",
    }
    messages = info["status"].map(reasons).fillna("Knorr Product_Line needs review")
    return is_review, SEV_YELLOW, messages


def _knorr_expected_breakout(df, ctx):
    n = len(df)
    brand = df[CFG["col_brand"]].astype("string").fillna("").str.strip()
    channel = df[CFG["col_channel"]].astype("string").fillna("").str.strip()
    is_scope = ((brand.str.casefold() == KNORR_BRAND.casefold())
                & (channel.str.casefold() == KNORR_SOCIAL_CHANNEL.casefold()))

    creative = df["Creative Name"].astype("string").fillna("") \
               if "Creative Name" in df.columns else pd.Series([""]*n, index=df.index)
    cr_cf = creative.str.casefold()
    has_cr = creative.str.strip() != ""

    is_other = cr_cf.apply(lambda s: any(k in s for k in KNORR_OTHER_SAY_KW))
    is_bsay  = cr_cf.apply(lambda s: any(k in s for k in KNORR_BRAND_SAY_KW))

    expected = pd.Series(pd.NA, index=df.index, dtype="string")
    status   = pd.Series("", index=df.index, dtype="string")

    m = is_scope & (~has_cr)
    status = status.mask(m, "review:no_creative")
    m = is_scope & has_cr & is_other & is_bsay
    status = status.mask(m, "review:ambiguous")
    m = is_scope & has_cr & is_other & (~is_bsay)
    expected = expected.mask(m, KNORR_OTHER_SAY_VAL); status = status.mask(m, "target")
    m = is_scope & has_cr & is_bsay & (~is_other)
    expected = expected.mask(m, KNORR_BRAND_SAY_VAL); status = status.mask(m, "target")
    m = is_scope & has_cr & (~is_other) & (~is_bsay)
    status = status.mask(m, "review:no_keyword")

    return pd.DataFrame({"expected": expected, "status": status}, index=df.index)


def rule_knorr_breakout_error(df, ctx):
    info = ctx["knorr_breakout"]
    actual = df[CFG["col_breakout"]].astype("string").fillna("").str.strip()
    has_target = info["status"] == "target"
    mismatch = (has_target
                & (actual.str.casefold() != info["expected"].str.casefold())).fillna(False)
    messages = ("Knorr/Social: Creative Name implies Breakout '"
                + info["expected"].astype(str).fillna("") + "' but Breakout is '" + actual + "'")
    return mismatch, SEV_RED, messages


def rule_knorr_breakout_review(df, ctx):
    info = ctx["knorr_breakout"]
    is_review = info["status"].str.startswith("review:").fillna(False)
    reasons = {
        "review:no_creative": "Knorr/Social: no Creative Name — can't determine Breakout",
        "review:ambiguous":   "Knorr/Social: Creative Name has BOTH Other Say and Brand Say keywords",
        "review:no_keyword":  "Knorr/Social: Creative Name has no Breakout keyword (BHAMP/IHAMP/BSAY)",
    }
    messages = info["status"].map(reasons).fillna("Knorr/Social Breakout needs review")
    return is_review, SEV_YELLOW, messages


RULES = (
    [_make_null_rule(c, except_tv=False) for c in CFG["critical_always"]] +
    [_make_null_rule(c, except_tv=True)  for c in CFG["critical_except_tv"]] +
    [
        rule_negative_cost,
        rule_social_channel_mismatch,
        rule_unexpected_structure,
        rule_placeholder_placement,
        rule_audience_mismatch,
        rule_audience_code_issue,
        rule_breakout_other_say,
        rule_partner_companion,
        rule_tv_audience,
        rule_channel_type_conflict,
        rule_twitter_x_corp,
        rule_knorr_pl_error,
        rule_knorr_pl_review,
        rule_knorr_breakout_error,
        rule_knorr_breakout_review,
    ]
)


# Rules whose flagged rows are summarized (count only, in "What to fix first" and
# the rule breakdown) but NOT listed row-by-row in the Review sheet — these are
# "fill in manually" cases (missing/blank required fields), not "go find and fix
# this specific value" cases like a Knorr mismatch or a Social code error.
ROW_DETAIL_EXCLUDE_PREFIXES = ("null__",)


# =============================================================================
# ENGINE (ROW-LEVEL)
# =============================================================================
def apply_rules(df, ctx):
    records, counts, severities = [], {}, {}
    for rule in RULES:
        mask, severity, messages = rule(df, ctx)
        mask = mask.fillna(False)
        n = int(mask.sum())
        counts[rule.__name__] = n
        severities[rule.__name__] = severity
        if n and not rule.__name__.startswith(ROW_DETAIL_EXCLUDE_PREFIXES):
            records.append(pd.DataFrame({
                "row": df.index[mask],
                "severity": severity,
                "message": messages[mask].values,
            }))

    if records:
        issues = pd.concat(records, ignore_index=True)
    else:
        issues = pd.DataFrame(columns=["row", "severity", "message"])

    if len(issues):
        issues["rank"] = issues["severity"].map({SEV_RED: 2, SEV_YELLOW: 1})
        agg = issues.groupby("row").agg(
            rank=("rank", "max"),
            detail=("message", lambda s: " | ".join(s)),
        )
        agg["severity"] = agg["rank"].map({2: SEV_RED, 1: SEV_YELLOW})
    else:
        agg = pd.DataFrame(columns=["rank", "detail", "severity"])
    return agg, counts, severities


# =============================================================================
# DELIVERY ANALYSIS: one tidy row per MONTH + UNIT
# =============================================================================
def _summarize_units(df, ctx):
    cost = ctx["cost_num"].fillna(0)
    impr = ctx["impr_num"].fillna(0)
    grps = ctx["grps_num"].fillna(0)
    channel = df[CFG["col_channel"]]
    team = (df[CFG["col_team"]].astype(str).fillna("") if CFG["col_team"] in df.columns
            else pd.Series([""] * len(df), index=df.index))
    is_tv = channel == CFG["tv_channel_value"]
    is_grp_off = channel.isin(CFG["grp_offline_channels"])
    is_impr_off = channel.isin(CFG["impr_offline_channels"])
    pn = df[CFG["col_placement"]]
    has_pn = pn.notna() & (pn.astype(str).fillna("").str.strip() != "")
    cols = ["Month", "Month_ts", "Unit_Type", "Unit", "Channel", "Brand", "Team",
            "Raw_Partner", "Total_Cost", "Total_Impressions", "Total_GRPs",
            "Expected_Metric"]
    frames = []

    digital = (~is_tv) & (~is_grp_off) & (~is_impr_off) & has_pn
    if digital.any():
        base = pd.DataFrame({
            "Month":    ctx["month_label"][digital].values,
            "Month_ts": ctx["month_ts"][digital].values,
            "key":      df.loc[digital, CFG["col_placement"]].astype(str).fillna("").values,
            "Channel":  channel[digital].astype(str).fillna("").values,
            "Brand":    df.loc[digital, CFG["col_brand"]].astype(str).fillna("").values,
            "Team":     team[digital].values,
            "partner":  df.loc[digital, CFG["col_raw_partner"]].values,
            "cost":     cost[digital].values,
            "impr":     impr[digital].values,
            "grps":     grps[digital].values,
        })
        g = base.groupby(["Month", "key"], as_index=False).agg(
            Month_ts=("Month_ts", "first"),
            Channel=("Channel", "first"),
            Brand=("Brand", "first"),
            Team=("Team", _join_distinct),
            Raw_Partner=("partner", _join_distinct),
            Total_Cost=("cost", "sum"),
            Total_Impressions=("impr", "sum"),
            Total_GRPs=("grps", "sum"),
        )
        g["Unit_Type"] = "Placement"
        g["Unit"] = g["key"]
        g["Expected_Metric"] = "Impressions"
        frames.append(g[cols])

    if is_tv.any():
        tv = pd.DataFrame({
            "Month":    ctx["month_label"][is_tv].values,
            "Month_ts": ctx["month_ts"][is_tv].values,
            "Brand":    df.loc[is_tv, CFG["col_brand"]].astype(str).fillna("").str.strip().values,
            "Team":     team[is_tv].values,
            "daypart":  df.loc[is_tv, CFG["col_daypart"]].astype(str).fillna("").str.strip().values,
            "network":  df.loc[is_tv, CFG["col_network"]].astype(str).fillna("").str.strip().values,
            "partner":  df.loc[is_tv, CFG["col_raw_partner"]].values,
            "cost":     cost[is_tv].values,
            "impr":     impr[is_tv].values,
            "grps":     grps[is_tv].values,
        })
        g = tv.groupby(["Month", "Brand", "daypart", "network"], as_index=False).agg(
            Month_ts=("Month_ts", "first"),
            Team=("Team", _join_distinct),
            Raw_Partner=("partner", _join_distinct),
            Total_Cost=("cost", "sum"),
            Total_Impressions=("impr", "sum"),
            Total_GRPs=("grps", "sum"),
        )
        g["Channel"] = CFG["tv_channel_value"]
        g["Unit_Type"] = "TV"
        g["Unit"] = g["daypart"] + " | " + g["network"]
        metric = g["Brand"].map(TV_METRIC_BY_BRAND)
        if TV_METRIC_DEFAULT is not None:
            metric = metric.fillna(TV_METRIC_DEFAULT)
        g["Expected_Metric"] = metric.fillna("UNMAPPED")
        frames.append(g[cols])

    if is_grp_off.any():
        off = pd.DataFrame({
            "Month":    ctx["month_label"][is_grp_off].values,
            "Month_ts": ctx["month_ts"][is_grp_off].values,
            "Channel":  channel[is_grp_off].astype(str).fillna("").values,
            "Brand":    df.loc[is_grp_off, CFG["col_brand"]].values,
            "Team":     team[is_grp_off].values,
            "partner":  df.loc[is_grp_off, CFG["col_raw_partner"]].astype(str).fillna("").str.strip().values,
            "cost":     cost[is_grp_off].values,
            "impr":     impr[is_grp_off].values,
            "grps":     grps[is_grp_off].values,
        })
        g = off.groupby(["Month", "Channel", "partner"], as_index=False).agg(
            Month_ts=("Month_ts", "first"),
            Brand=("Brand", _join_distinct),
            Team=("Team", _join_distinct),
            Total_Cost=("cost", "sum"),
            Total_Impressions=("impr", "sum"),
            Total_GRPs=("grps", "sum"),
        )
        g["Unit_Type"] = "Offline (Raw_Partner)"
        g["Unit"] = g["partner"]
        g["Raw_Partner"] = g["partner"]
        g["Expected_Metric"] = "GRPs"
        frames.append(g[cols])

    if is_impr_off.any():
        off = pd.DataFrame({
            "Month":    ctx["month_label"][is_impr_off].values,
            "Month_ts": ctx["month_ts"][is_impr_off].values,
            "Channel":  channel[is_impr_off].astype(str).fillna("").values,
            "Brand":    df.loc[is_impr_off, CFG["col_brand"]].values,
            "Team":     team[is_impr_off].values,
            "partner":  df.loc[is_impr_off, CFG["col_raw_partner"]].astype(str).fillna("").str.strip().values,
            "cost":     cost[is_impr_off].values,
            "impr":     impr[is_impr_off].values,
            "grps":     grps[is_impr_off].values,
        })
        g = off.groupby(["Month", "Channel", "partner"], as_index=False).agg(
            Month_ts=("Month_ts", "first"),
            Brand=("Brand", _join_distinct),
            Team=("Team", _join_distinct),
            Total_Cost=("cost", "sum"),
            Total_Impressions=("impr", "sum"),
            Total_GRPs=("grps", "sum"),
        )
        g["Unit_Type"] = "Offline (Raw_Partner)"
        g["Unit"] = g["partner"]
        g["Raw_Partner"] = g["partner"]
        g["Expected_Metric"] = "Impressions"
        frames.append(g[cols])

    if not frames:
        return pd.DataFrame(columns=cols)
    return pd.concat(frames, ignore_index=True)


def build_delivery_issues(df, ctx):
    units = _summarize_units(df, ctx)
    if units.empty:
        return pd.DataFrame(columns=DELIVERY_COLS)
    units = units.reset_index(drop=True)
    units["uid"] = units.index

    c = units["Total_Cost"]
    i = units["Total_Impressions"]
    g = units["Total_GRPs"]
    m = units["Expected_Metric"]
    is_grp, is_impr, is_unm = (m == "GRPs"), (m == "Impressions"), (m == "UNMAPPED")

    if CROSS_METRIC_INCLUDE_TV:
        cross_ok = pd.Series(True, index=units.index)
    else:
        cross_ok = units["Unit_Type"] != "TV"

    checks = [
        (is_grp & (c > 0) & (g == 0), "ERROR",
         "Measured by GRPs: spend but 0 GRPs in the month"),
        (is_impr & (c > 0) & (i == 0), "ERROR",
         "Measured by Impressions: spend but 0 impressions in the month"),
        (is_grp & (g > 0) & (c == 0), "REVIEW",
         "Measured by GRPs: GRPs but 0 spend (possible Added Value)"),
        (is_impr & (i > 0) & (c == 0), "REVIEW",
         "Measured by Impressions: impressions but 0 spend (possible Added Value)"),
        (is_grp & (i > 0) & cross_ok, METRIC_CROSS_SEVERITY,
         "Measured by GRPs but has impressions (expected 0/empty)"),
        (is_impr & (g > 0) & cross_ok, METRIC_CROSS_SEVERITY,
         "Measured by Impressions but has GRPs (expected 0/empty)"),
        (is_unm & (c > 0), "REVIEW",
         "TV brand has no defined delivery metric; add it to TV_METRIC_BY_BRAND"),
    ]

    records = []
    for mask, sev, msg in checks:
        mask = mask.fillna(False)
        if mask.any():
            sub = units[mask].copy()
            sub["Severity"] = sev
            sub["Issue"] = msg
            records.append(sub)

    with np.errstate(divide="ignore", invalid="ignore"):
        cpm_arr = np.where(i.values > 0, c.values / i.values * 1000.0, 0.0)
    cpm = pd.Series(cpm_arr, index=units.index)
    in_scope = units["Unit_Type"].isin(CPM_CHECK_UNIT_TYPES)
    high_cpm = (in_scope & (i > 0) & (c > 0)
                & (cpm > CPM_HIGH_THRESHOLD)).fillna(False)
    if high_cpm.any():
        sub = units[high_cpm].copy()
        sub["Severity"] = "ERROR" if CPM_HIGH_SEVERITY == "ERROR" else "REVIEW"
        sub["Issue"] = ("High CPM: $" + cpm[high_cpm].round(2).astype(str).fillna("")
                        + f" per 1,000 impressions (threshold ${CPM_HIGH_THRESHOLD:.0f})")
        records.append(sub)

    if not records:
        return pd.DataFrame(columns=DELIVERY_COLS)

    allrec = pd.concat(records, ignore_index=True)
    allrec["rank"] = allrec["Severity"].map({"ERROR": 2, "REVIEW": 1})
    agg = allrec.groupby("uid", as_index=False).agg(
        rank=("rank", "max"),
        Issue=("Issue", lambda s: " | ".join(s)),
        Channel=("Channel", "first"),
        Brand=("Brand", "first"),
        Team=("Team", "first"),
        Raw_Partner=("Raw_Partner", "first"),
        Unit_Type=("Unit_Type", "first"),
        Unit=("Unit", "first"),
        Month=("Month", "first"),
        Month_ts=("Month_ts", "first"),
        Total_Impressions=("Total_Impressions", "first"),
        Total_GRPs=("Total_GRPs", "first"),
        Total_Cost=("Total_Cost", "first"),
    )
    agg["Severity"] = agg["rank"].map({2: "ERROR", 1: "REVIEW"})
    agg = agg.rename(columns={"Total_Impressions": "Total Impressions",
                              "Total_GRPs": "Total GRPs",
                              "Total_Cost": "Total Cost"})
    agg["_order"] = np.where(agg["Severity"] == "ERROR", 0, 1)
    agg = agg.sort_values(["_order", "Month_ts", "Total Cost"],
                          ascending=[True, True, False], na_position="last")
    return agg[DELIVERY_COLS].reset_index(drop=True)


# =============================================================================
# SUMMARY BUILDERS: channel-by-month, offline presence, reconciliation
# =============================================================================
def _derive_master_channel(channel_series: pd.Series) -> pd.Series:
    ch = channel_series.astype(str).fillna("").str.strip()
    return ch.where(~ch.isin(DIGITAL_CHANNELS), MASTER_CHANNEL_DIGITAL_LABEL)


def build_channel_summary(df, ctx):
    cm = pd.DataFrame({
        "Month": ctx["month_label"].values,
        "Month_ts": ctx["month_ts"].values,
        "Channel": df[CFG["col_channel"]].astype(str).fillna("").values,
        "Cost": ctx["cost_num"].fillna(0).values,
        "Impr": ctx["impr_num"].fillna(0).values,
        "GRPs": ctx["grps_num"].fillna(0).values,
    })
    g = cm.groupby(["Month", "Channel"], as_index=False).agg(
        Month_ts=("Month_ts", "first"),
        Rows=("Cost", "size"),
        Total_Cost=("Cost", "sum"),
        Total_Impressions=("Impr", "sum"),
        Total_GRPs=("GRPs", "sum"),
    )
    g = (g.sort_values(["Month_ts", "Channel"], na_position="last")
           .drop(columns="Month_ts"))
    g = g.rename(columns={"Rows": "# Rows", "Total_Cost": "Total Cost",
                          "Total_Impressions": "Total Impressions",
                          "Total_GRPs": "Total GRPs"})
    return g[["Month", "Channel", "# Rows", "Total Cost",
              "Total Impressions", "Total GRPs"]]


def build_offline_presence(df, ctx):
    present = pd.DataFrame({
        "Month": ctx["month_label"].values,
        "Month_ts": ctx["month_ts"].values,
        "Channel": df[CFG["col_channel"]].astype(str).fillna("").values,
    })
    months = (present[["Month", "Month_ts"]].drop_duplicates("Month")
              .sort_values("Month_ts", na_position="last"))
    rows = []
    for _, r in months.iterrows():
        mlabel = r["Month"]
        chans = set(present.loc[present["Month"] == mlabel, "Channel"])
        missing = [c for c in CFG["expected_offline_channels"] if c not in chans]
        status = "All present" if not missing else "Missing: " + ", ".join(missing)
        rows.append({"Month": mlabel, "Offline/Traditional-non TV": status})
    return pd.DataFrame(rows, columns=["Month", "Offline/Traditional-non TV"])


def build_reconciliation(df, ctx):
    channel = df[CFG["col_channel"]]
    is_tv = channel == CFG["tv_channel_value"]
    is_grp_off = channel.isin(CFG["grp_offline_channels"])
    is_impr_off = channel.isin(CFG["impr_offline_channels"])
    pn = df[CFG["col_placement"]]
    has_pn = pn.notna() & (pn.astype(str).fillna("").str.strip() != "")
    is_place = (~is_tv) & (~is_grp_off) & (~is_impr_off) & has_pn
    is_excl = (~is_tv) & (~is_grp_off) & (~is_impr_off) & (~has_pn)
    total = len(df)
    a = int(is_tv.sum())
    b = int(is_grp_off.sum())
    e = int(is_impr_off.sum())
    c = int(is_place.sum())
    d = int(is_excl.sum())
    ok = (a + b + e + c + d) == total
    recon = pd.DataFrame({
        "Check": [
            "Total rows read",
            "Evaluated by row-level rules",
            "-- Delivery analysis coverage --",
            "TV rows (Brand+Daypart+Network)",
            "Print rows (Raw_Partner, GRPs)",
            "OOH/DOOH/Cinema rows (Raw_Partner, Impressions)",
            "Placement rows (Package_Placement_Name, incl. Radio)",
            "Not included (no placement, not TV/offline)",
            "All rows accounted for",
        ],
        "Value": [
            f"{total:,}",
            f"{total:,} (100%)",
            "",
            f"{a:,}",
            f"{b:,}",
            f"{e:,}",
            f"{c:,}",
            f"{d:,}",
            "YES" if ok else "NO - CHECK",
        ],
    })
    return recon


def _friendly_rule_name(rule_key: str) -> str:
    if rule_key.startswith("null__"):
        return f"Missing: {rule_key[len('null__'):]}"
    return RULE_LABELS.get(rule_key, rule_key)


def build_verdict(agg, delivery, counts, total_rows,
                  n_red, n_yellow, n_d_red, n_d_yellow):
    n_flagged = len(agg)
    n_deliv = len(delivery)
    pct = (n_flagged / total_rows) if total_rows else 0.0

    if n_red > 0 or n_d_red > 0:
        status, level = "ACTION REQUIRED", SEV_RED
    elif n_yellow > 0 or n_d_yellow > 0:
        status, level = "REVIEW SUGGESTED", SEV_YELLOW
    else:
        status, level = "ALL CLEAR", "green"

    # "Top cause" is picked among the rules that actually show up in Review —
    # missing-field rules (null__...) are summarized by count elsewhere (see
    # ROW_DETAIL_EXCLUDE_PREFIXES) and would make a nonsensical % of n_flagged
    # here, since n_flagged no longer includes rows flagged only by those.
    fired_detail = {k: v for k, v in counts.items()
                    if v > 0 and not k.startswith(ROW_DETAIL_EXCLUDE_PREFIXES)}
    n_missing_fields = sum(v for k, v in counts.items()
                           if v > 0 and k.startswith(ROW_DETAIL_EXCLUDE_PREFIXES))
    if fired_detail:
        top_key = max(fired_detail, key=lambda k: fired_detail[k])
        top_cause = f"{_friendly_rule_name(top_key)} ({fired_detail[top_key]:,})"
    elif n_missing_fields:
        top_cause = f"Missing required fields ({n_missing_fields:,} — see 'What to fix first')"
    else:
        top_cause = "None"

    verdict = pd.DataFrame({
        "Field": ["Status", "Rows requiring attention",
                  "Delivery incidences", "Top cause"],
        "Value": [status,
                  f"{n_flagged:,} of {total_rows:,} ({pct:.1%})",
                  f"{n_deliv:,}",
                  top_cause],
    })

    comment = _bruno_excel_comment(n_flagged, total_rows, fired_detail, n_deliv, n_missing_fields)
    return verdict, level, comment


def _categorize_delivery(part, fallback_sev):
    pl = part.lower()
    for cat, keys, sev in _DELIVERY_CATS:
        if any(k in pl for k in keys):
            return cat, sev
    return "Other delivery issue", fallback_sev


def build_priority_issues(counts, severities, delivery):
    rows = []
    for name, n in counts.items():
        if n > 0:
            sev = "ERROR" if severities.get(name) == SEV_RED else "REVIEW"
            rows.append({"Issue": _friendly_rule_name(name),
                         "Scope": f"{n:,} rows", "Severity": sev, "_n": n})

    if delivery is not None and len(delivery):
        dc = {}
        for _, dr in delivery.iterrows():
            for part in str(dr["Issue"]).split(" | "):
                cat, sev = _categorize_delivery(part, dr["Severity"])
                dc[(cat, sev)] = dc.get((cat, sev), 0) + 1
        for (cat, sev), n in dc.items():
            rows.append({"Issue": cat, "Scope": f"{n:,} units",
                         "Severity": sev, "_n": n})

    frame = pd.DataFrame(rows, columns=["Issue", "Scope", "Severity", "_n"])
    if len(frame):
        frame["_o"] = frame["Severity"].map({"ERROR": 0, "REVIEW": 1})
        frame = (frame.sort_values(["_o", "_n"], ascending=[True, False])
                      .drop(columns=["_o", "_n"]).reset_index(drop=True))
        n_passed = sum(1 for v in counts.values() if v == 0)
        note = f"{n_passed} row-level rule(s) passed with no incidences"
    else:
        frame = frame.drop(columns="_n")
        note = "All checks passed — nothing to fix"
    return frame, note


def build_checklist(df, ctx):
    work = pd.DataFrame({
        "Brand":          df[CFG["col_brand"]].astype(str).fillna("").str.strip(),
        "Category":       df[CFG["col_category"]].astype(str).fillna("").str.strip(),
        "Master Channel": _derive_master_channel(df[CFG["col_channel"]]),
        "_date":          ctx["dates"],
    })
    g = work.groupby(["Brand", "Category", "Master Channel"],
                     dropna=False, as_index=False).agg(
        earliest=("_date", "min"),
        latest=("_date", "max"),
    )

    def _fmt(ts):
        return ts.strftime("%Y-%m-%d") if pd.notna(ts) else "No valid date"

    g["Earliest Data Delivered"] = g["earliest"].map(_fmt)
    g["Latest Data Delivered"]   = g["latest"].map(_fmt)
    g = (g.drop(columns=["earliest", "latest"])
           .sort_values(["Brand", "Category", "Master Channel"])
           .reset_index(drop=True))
    return g[["Brand", "Category", "Master Channel",
              "Earliest Data Delivered", "Latest Data Delivered"]]


def _bruno_excel_comment(n_flagged, total, fired, n_delivery, n_missing_fields=0):
    pct = n_flagged / total if total else 0
    missing_note = (f" Separately, {n_missing_fields:,} row(s) have a missing "
                    f"required field — see 'What to fix first'." if n_missing_fields else "")

    if n_flagged == 0 and n_delivery == 0:
        if n_missing_fields:
            return (f"No values-to-correct found — but {n_missing_fields:,} row(s) "
                    f"have a missing required field to fill in. See 'What to fix first'.")
        return "Every rule passed. Clean data — nothing to flag this month."

    if n_flagged == 0 and n_delivery > 0:
        return (f"Row-level checks came back clean — but {n_delivery} "
                f"delivery issue(s) need attention. Check Delivery_Issues."
                + missing_note)

    if fired:
        top_k = max(fired, key=lambda k: fired[k])
        top_v = fired[top_k]
        top_pct = top_v / n_flagged
        top_name = _friendly_rule_name(top_k)

        if top_pct >= 0.90 and len(fired) == 1:
            return (f"{top_v:,} of {n_flagged:,} flags ({top_pct:.0%}) trace back "
                    f"to one field: {top_name}. Fix that and this data is "
                    f"essentially clean." + missing_note)
        if top_pct >= 0.90 and len(fired) > 1:
            others = len(fired) - 1
            return (f"{top_name} accounts for {top_pct:.0%} of all flags. "
                    f"{others} other rule(s) fired but are marginal. Start there."
                    + missing_note)
        if pct < 0.02:
            return (f"{n_flagged:,} flags across {len(fired)} rules — only "
                    f"{pct:.1%} of total rows. Contained. Lead with {top_name} "
                    f"({top_v:,} rows)." + missing_note)
        return (f"{n_flagged:,} flags across {len(fired)} rules ({pct:.1%}). "
                f"Several areas need attention. Start with {top_name} "
                f"({top_v:,} rows)." + missing_note)

    if n_missing_fields:
        return (f"No values-to-correct beyond missing fields — {n_missing_fields:,} "
                f"row(s) need a required field filled in. See 'What to fix first'.")
    return "Analysis complete. Review the sections below."


def build_output_filename(df, ctx):
    brands = sorted({str(b).strip() for b in df[CFG["col_brand"]].dropna()
                     if str(b).strip() != ""})
    if len(brands) == 1:
        brand_part = brands[0]
    elif len(brands) == 0:
        brand_part = "NoBrand"
    else:
        brand_part = "Multi-Brand"
    valid = ctx["dates"].dropna()
    if len(valid):
        start, end = valid.min().to_period("M"), valid.max().to_period("M")
        span = (_period_to_label(start) if start == end
                else f"{_period_to_label(start)} - {_period_to_label(end)}")
    else:
        span = "Unknown period"
    return _sanitize_filename(f"{OUTPUT_PREFIX} {brand_part} {span}.xlsx")


# =============================================================================
# WRITE EXCEL — to BytesIO instead of a disk path
# =============================================================================
def write_excel_bytes(df, agg, counts, severities, delivery, channel_summary, offline_presence,
                       recon, checklist, input_filename, months_str, brands_str, total_rows) -> bytes:
    if len(agg):
        review = df.loc[agg.index].copy()
        label = agg["severity"].map({SEV_RED: "ERROR", SEV_YELLOW: "REVIEW"})
        review.insert(0, "Severity", label.values)
        review.insert(1, "Review_Detail", agg["detail"].values)
        review.insert(2, "Original_Excel_Row", (agg.index + 2))
        review["_order"] = agg["severity"].map({SEV_RED: 0, SEV_YELLOW: 1}).values
        review = review.sort_values("_order").drop(columns="_order")
    else:
        review = pd.DataFrame(columns=["Severity", "Review_Detail",
                                       "Original_Excel_Row"] + list(df.columns))

    n_red    = int((agg["severity"] == SEV_RED).sum())    if len(agg) else 0
    n_yellow = int((agg["severity"] == SEV_YELLOW).sum()) if len(agg) else 0
    n_d_red    = int((delivery["Severity"] == "ERROR").sum())  if len(delivery) else 0
    n_d_yellow = int((delivery["Severity"] == "REVIEW").sum()) if len(delivery) else 0

    verdict, verdict_level, bruno_comment = build_verdict(
        agg, delivery, counts, total_rows, n_red, n_yellow, n_d_red, n_d_yellow)
    priority, priority_note = build_priority_issues(counts, severities, delivery)

    meta = pd.DataFrame({
        "Field": ["Months in data", "Brand(s)", "Input file", "Run timestamp",
                  "Total rows", "Rows flagged", "Errors (red)", "Review (yellow)",
                  "Delivery issues", "Delivery errors (red)", "Delivery review (yellow)"],
        "Value": [months_str, brands_str, input_filename,
                  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                  f"{total_rows:,}", f"{len(agg):,}", f"{n_red:,}", f"{n_yellow:,}",
                  f"{len(delivery):,}", f"{n_d_red:,}", f"{n_d_yellow:,}"],
    })

    breakdown = (pd.DataFrame({"Rule": [_friendly_rule_name(k) for k in counts],
                               "Flagged rows": list(counts.values())})
                 .sort_values("Flagged rows", ascending=False))

    grps_zero = (channel_summary["Total GRPs"].fillna(0).eq(0).all()
                 if len(channel_summary) else False)
    chan_note = ("Note: Total GRPs = 0 is expected when TV is measured by "
                 "Impressions for this brand (see TV_METRIC_BY_BRAND)."
                 if grps_zero else None)
    off_note = "If these channels are not bought for this brand, this reminder is expected."

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_blocks = []
        cursor = 0

        def add_block(title, frame, note=None, extra_rows=0):
            nonlocal cursor
            title_row = cursor
            header_startrow = cursor + 1
            frame.to_excel(writer, sheet_name="Summary", index=False,
                           startrow=header_startrow)
            summary_blocks.append({
                "title": title, "title_row": title_row,
                "header_startrow": header_startrow,
                "nrows": len(frame), "ncols": frame.shape[1], "note": note,
            })
            cursor = (header_startrow + 1 + len(frame)
                      + extra_rows + (1 if note else 0) + 2)

        add_block(VERDICT_TITLE, verdict, extra_rows=1)
        add_block(PRIORITY_TITLE, priority, note=priority_note)
        add_block("Run information", meta)
        add_block(CHANNEL_TITLE, channel_summary, note=chan_note)
        add_block("Offline/Traditional-non TV presence (by month)",
                  offline_presence, note=off_note)
        add_block("Coverage check (audit)", recon)
        add_block("All rules checked (audit)", breakdown)

        review.to_excel(writer, sheet_name="Review", index=False, startrow=1)
        delivery.to_excel(writer, sheet_name="Delivery_Issues", index=False)
        checklist.to_excel(writer, sheet_name="AA Checklist", index=False)

        ws_s = writer.sheets["Summary"]
        level_fill = {SEV_RED: FILL_RED, SEV_YELLOW: FILL_YELLOW, "green": FILL_GREEN}

        for b in summary_blocks:
            title, ncols, nrows = b["title"], b["ncols"], b["nrows"]
            ws_s.cell(row=b["title_row"] + 1, column=1,
                      value=title).font = Font(bold=True, size=12)

            hr = b["header_startrow"] + 1
            for cidx in range(1, ncols + 1):
                cell = ws_s.cell(row=hr, column=cidx)
                cell.font = Font(bold=True)
                cell.fill = FILL_HEADER

            first_data, last_data = hr + 1, hr + nrows

            if title == VERDICT_TITLE:
                fill = level_fill.get(verdict_level, FILL_YELLOW)
                for r in range(first_data, last_data + 1):
                    for c in range(1, ncols + 1):
                        ws_s.cell(row=r, column=c).fill = fill
                ws_s.cell(row=first_data, column=1).font = Font(bold=True)
                ws_s.cell(row=first_data, column=2).font = Font(bold=True, size=12)

                comment_row = last_data + 1
                ws_s.merge_cells(start_row=comment_row, start_column=1,
                                 end_row=comment_row, end_column=4)
                cc = ws_s.cell(row=comment_row, column=1, value=bruno_comment)
                cc.font = Font(italic=True, size=10, color="444444")
                cc.fill = PatternFill("solid", fgColor="FAFAFA")
                cc.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
                ws_s.row_dimensions[comment_row].height = 30
            if title == PRIORITY_TITLE and nrows > 0:
                for r in range(first_data, last_data + 1):
                    sev_val = ws_s.cell(row=r, column=3).value
                    fill = FILL_RED if sev_val == "ERROR" else FILL_YELLOW
                    for c in range(1, ncols + 1):
                        ws_s.cell(row=r, column=c).fill = fill

            if title == CHANNEL_TITLE:
                for r in range(first_data, last_data + 1):
                    ws_s.cell(row=r, column=3).number_format = "#,##0"
                    ws_s.cell(row=r, column=4).number_format = "#,##0.00"
                    ws_s.cell(row=r, column=5).number_format = "#,##0"
                    ws_s.cell(row=r, column=6).number_format = "#,##0"

            if b["note"]:
                nc = ws_s.cell(row=last_data + 1, column=1, value=b["note"])
                nc.font = Font(italic=True, size=9)

        for col, w in {"A": 40, "B": 46, "C": 16, "D": 18, "E": 18, "F": 14}.items():
            ws_s.column_dimensions[col].width = w

        ws = writer.sheets["Review"]
        ws.merge_cells(start_row=1, start_column=1, end_row=1,
                       end_column=max(review.shape[1], 4))
        note_cell = ws.cell(row=1, column=1, value=(
            "Missing/blank required fields are NOT listed here row-by-row (they're "
            "filled in manually) — see their counts in 'What to fix first' and "
            "'All rules checked' on the Summary tab instead."
        ))
        note_cell.font = Font(italic=True, size=9, color="666666")
        ws.row_dimensions[1].height = 18
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.fill = FILL_HEADER
        ws.freeze_panes = "A3"
        if ws.max_row >= 2 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A2:{ws.cell(row=2, column=ws.max_column).coordinate}"
        if ws.max_row >= 3:
            # iter_rows() iterates the sheet's internal storage directly; looping
            # with ws.cell(row, column) instead (as this used to) is ~300x slower
            # at this scale — that repeated-lookup cost, not the coloring itself,
            # was the real reason "Writing Excel report" felt slow.
            for row in ws.iter_rows(min_row=3):
                fill = FILL_RED if row[0].value == "ERROR" else FILL_YELLOW
                for cell in row:
                    cell.fill = fill
        ws.column_dimensions["B"].width = 70

        wd = writer.sheets["Delivery_Issues"]
        for cell in wd[1]:
            cell.font = Font(bold=True)
            cell.fill = FILL_HEADER
        wd.freeze_panes = "A2"
        if wd.max_row >= 1 and wd.max_column >= 1:
            wd.auto_filter.ref = wd.dimensions
        if wd.max_row >= 2:
            for row in wd.iter_rows(min_row=2):
                fill = FILL_RED if row[0].value == "ERROR" else FILL_YELLOW
                for cell in row:
                    cell.fill = fill
                row[9].number_format = "#,##0"       # Total Impressions
                row[10].number_format = "#,##0"      # Total GRPs
                row[11].number_format = "#,##0.00"   # Total Cost
        widths = {"A": 10, "B": 60, "C": 14, "D": 18, "E": 16, "F": 24,
                  "G": 22, "H": 46, "I": 14, "J": 18, "K": 12, "L": 16}
        for col, w in widths.items():
            wd.column_dimensions[col].width = w

        wc = writer.sheets["AA Checklist"]
        wc.sheet_properties.tabColor = "FF8357"  # coral
        for cell in wc[1]:
            cell.font = Font(bold=True)
            cell.fill = FILL_HEADER
        wc.freeze_panes = "A2"
        if wc.max_row >= 1 and wc.max_column >= 1:
            wc.auto_filter.ref = wc.dimensions
        for col, w in {"A": 22, "B": 22, "C": 24, "D": 24, "E": 24}.items():
            wc.column_dimensions[col].width = w

    return output.getvalue(), verdict, verdict_level, bruno_comment, priority, priority_note, \
        n_red, n_yellow, n_d_red, n_d_yellow


# ----------------------------------------------------------------------------
# In-app styling for the "What to fix first" table
# ----------------------------------------------------------------------------
def _style_priority(priority: pd.DataFrame):
    def highlight(row):
        color = FILL_RED if row["Severity"] == "ERROR" else FILL_YELLOW
        return [f"background-color: #{str(color.fgColor.rgb)[-6:]}; color: #1a1a1a"] * len(row)
    return priority.style.apply(highlight, axis=1)


def _style_offline_presence(offline_presence: pd.DataFrame):
    def highlight(row):
        col = "Offline/Traditional-non TV"
        color = "" if row[col] == "All present" else f"background-color: #{str(FILL_YELLOW.fgColor.rgb)[-6:]}; color: #1a1a1a"
        return [color] * len(row)
    return offline_presence.style.apply(highlight, axis=1)


def _build_rule_chart(counts: dict, severities: dict):
    """Horizontal bar chart of rules that fired at least once, colored by
    severity — the same information as the 'All rules checked' breakdown in
    the Excel report, but visual. Uses Altair (bundled with Streamlit)."""
    import altair as alt

    rows = [{"Rule": _friendly_rule_name(k), "Flagged rows": v,
             "Severity": "ERROR" if severities.get(k) == SEV_RED else "REVIEW"}
            for k, v in counts.items() if v > 0]
    if not rows:
        return None
    chart_df = pd.DataFrame(rows).sort_values("Flagged rows", ascending=True)

    return (
        alt.Chart(chart_df)
        .mark_bar()
        .encode(
            x=alt.X("Flagged rows:Q", axis=alt.Axis(format="d", tickMinStep=1)),
            y=alt.Y("Rule:N", sort=None, title=None),
            color=alt.Color(
                "Severity:N",
                scale=alt.Scale(domain=["ERROR", "REVIEW"], range=["#E25B45", "#FAC172"]),
                legend=alt.Legend(title=None),
            ),
            tooltip=["Rule", "Flagged rows", "Severity"],
        )
        .properties(height=max(24 * len(chart_df), 120))
    )


# ----------------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------------
CHECKS_REFERENCE_MD = """
### 1. Missing required fields (counted, not listed row-by-row)
These are summarized only — they're filled in manually, not values to go look up.

| Field | Severity | Exempt on TV rows |
|---|---|---|
| Channel | ERROR | No |
| Brand | ERROR | No |
| Product_Line | ERROR | No |
| Category | ERROR | No |
| Raw_Partner | ERROR | No |
| Audience | ERROR | No |
| Retailer | ERROR | No |
| Breakout | ERROR | No |
| Campaign | ERROR | Yes |
| Prisma_Campaign_Secondary | ERROR | Yes |
| Package_Placement_Name | ERROR | Yes |

### 2. Value-level checks (listed row-by-row in Review)

| Check | What it catches | Severity |
|---|---|---|
| Negative cost | Media_Cost is negative | ERROR |
| Social channel mismatch | The placement's inventory code implies Social but Channel isn't "Digital Social" (or vice versa) | REVIEW |
| Unexpected placement structure | The Placement Name doesn't match the standardized block (method + buy type + audience...); channel/audience couldn't be derived | REVIEW |
| Placeholder placement | Placement Name contains "Dummy", "DELETE", "DO NOT USE", or "DNU" — possibly a row to remove | REVIEW |
| Audience mismatch | The placement's audience code, per the catalog, doesn't match the Audience column | ERROR |
| Audience code issue | The audience code isn't in the catalog, or isn't a 4-digit code | REVIEW |
| 'Other Say' outside Social | Breakout = "Other Say" on a non-Digital-Social channel | ERROR |
| Partner missing companion | Channel = "Digital FEP" with Raw_Partner = "THE TRADE DESK INC" missing its expected companion (e.g. "THE TRADE DESK INC - X") | REVIEW |
| TV audience incorrect | On TV, Audience doesn't match what's expected for that brand (e.g. Hellmann's → A2564, Knorr → P2+) | ERROR |
| Channel/format conflict | Placement Name mixes "Audio" and "Video" in the same Online_X_Online_Y block | ERROR |
| X Corp partner missing | Inventory code is Twitter/X but Raw_Partner doesn't end in "- X CORP" | ERROR |
| Knorr Product_Line incorrect | Based on Prisma (Social), Placement Name (non-Social), or Daypart (TV), Product_Line should be Bouillon or Sides and isn't, or isn't either | ERROR |
| Knorr Product_Line — needs review | Knorr/Social row with no Creative Name, or ambiguous/missing keywords — can't determine Bouillon vs. Sides | REVIEW |
| Knorr Breakout incorrect | On Knorr + Digital Social, Creative Name implies a different Breakout ("Other Say" or "Brand Say") than the row has | ERROR |
| Knorr Breakout — needs review | Same case but no Creative Name, or ambiguous/missing keywords | REVIEW |

### 3. Delivery vs. spend (per unit: placement, TV, or offline partner — aggregated by month)
Each unit is measured against one expected metric (Impressions for digital/OOH/Cinema, GRPs for Print, and for TV it depends on the brand — see `TV_METRIC_BY_BRAND`).

| Check | What it catches | Severity |
|---|---|---|
| Spend but 0 GRPs | GRP-measured unit had spend but zero GRPs for the month | ERROR |
| Spend but 0 Impressions | Impression-measured unit had spend but zero impressions for the month | ERROR |
| GRPs but 0 spend | GRPs with no spend — possible Added Value | REVIEW |
| Impressions but 0 spend | Impressions with no spend — possible Added Value | REVIEW |
| Metric cross-contamination (GRPs) | GRP-measured unit also has impressions (expected 0/empty) | REVIEW |
| Metric cross-contamination (Impressions) | Impression-measured unit also has GRPs (expected 0/empty) | REVIEW |
| Unmapped TV metric | TV brand has no metric defined in `TV_METRIC_BY_BRAND` | REVIEW |
| High CPM | CPM above $50 (Placement, TV, and Offline-by-Raw_Partner units only) | REVIEW |

### 4. Informational summaries (not pass/fail checks)
- **Channel summary** — rows, cost, impressions, and GRPs by Channel and month.
- **Offline/Traditional-non TV presence** — whether Print/OOH/DOOH/Cinema all showed up each month, or which are missing (a "Missing" flag is expected if a brand doesn't buy those channels).
- **Coverage check (audit)** — how many rows were covered by each delivery-analysis path (TV, Print, OOH/DOOH/Cinema, Placement) vs. the file's total, to confirm no rows silently fell through the cracks.
- **Checklist** — earliest and latest data date delivered, by Brand + Category + Master Channel, for that month.

### How the verdict is built
- **ACTION REQUIRED** (red): at least one ERROR, from either row-level or delivery checks.
- **REVIEW SUGGESTED** (amber): no ERRORs, but at least one REVIEW.
- **ALL CLEAR** (green): nothing fired.

The "Top cause" and verdict comment only consider sections 2 and 3 (the row-by-row checks) — missing fields from section 1 are called out separately, with their own count.
"""


def render():
    if st.button("← Back to menu", key="inspect_back"):
        reset_all()
        st.session_state.active_tool = None
        st.rerun()

    init_state()

    st.title("Merit Inspect")
    st.caption(
        "Runs the monthly QA pass on a raw delivery file: row-level rule checks "
        "(nulls, Knorr Product_Line/Breakout, TV audience, Twitter/X partner, "
        "placeholders...), a spend-vs-delivery analysis per unit, and a "
        "brand/category/channel coverage checklist."
    )

    with st.expander("What checks does this run?"):
        st.markdown(CHECKS_REFERENCE_MD)

    uploader_key = st.session_state["inspect_uploader_key"]
    uploaded = st.file_uploader(
        "Upload the raw Excel file", type=["xlsx", "xls"],
        key=f"inspect_uploader_{uploader_key}",
    )

    with st.expander("Advanced: use a different audience catalog"):
        st.caption(
            "Leave this empty to use the built-in tools/config/audience_codes.csv. "
            "Only override it if you have a different approved code list (columns: "
            "Code, Audience)."
        )
        custom_catalog = st.file_uploader(
            "Audience catalog (.csv)", type=["csv"],
            key=f"inspect_catalog_uploader_{uploader_key}",
        )

    if uploaded and st.button("Run QA checks", type="primary", use_container_width=True):
        try:
            with st.status("Running QA checks...", expanded=True) as status:
                t0 = time.perf_counter()
                status.write("Loading file...")
                df = pd.read_excel(uploaded, sheet_name=0, engine="openpyxl").reset_index(drop=True)
                status.write(f"{len(df):,} rows loaded in {time.perf_counter() - t0:.1f}s.")

                required = set(CFG["critical_always"] + CFG["critical_except_tv"]
                               + [CFG["col_cost"], CFG["col_impressions"], CFG["col_grps"],
                                  CFG["col_daypart"], CFG["col_network"], CFG["col_date"]])
                missing = sorted(required - set(df.columns))
                if missing:
                    raise ValueError("Missing expected columns: " + ", ".join(missing))

                catalog_bytes = custom_catalog.getvalue() if custom_catalog is not None else None
                catalog = load_audience_catalog(catalog_bytes)
                catalog_warning = (None if catalog is not None else
                                   "No audience catalog found — audience checks skipped. "
                                   "Add tools/config/audience_codes.csv, or upload one above.")

                status.write("Running row-level rules...")
                t1 = time.perf_counter()
                ctx = build_context(df, catalog)
                agg, counts, severities = apply_rules(df, ctx)
                status.write(f"{len(RULES)} rules applied in {time.perf_counter() - t1:.1f}s.")

                status.write("Analyzing delivery vs spend...")
                t2 = time.perf_counter()
                delivery = build_delivery_issues(df, ctx)
                status.write(f"Delivery analysis done in {time.perf_counter() - t2:.1f}s.")

                status.write("Building summary tables...")
                channel_summary = build_channel_summary(df, ctx)
                offline_presence = build_offline_presence(df, ctx)
                recon = build_reconciliation(df, ctx)
                checklist = build_checklist(df, ctx)

                periods = pd.Series(ctx["month_period"].dropna().unique()).sort_values()
                months_str = (", ".join(_period_to_label(p) for p in periods)
                              if len(periods) else "Unknown")
                brands = sorted({str(b).strip() for b in df[CFG["col_brand"]].dropna()
                                 if str(b).strip()})
                brands_str = ", ".join(brands) if brands else "None"

                status.write("Writing Excel report...")
                t3 = time.perf_counter()
                (excel_bytes, verdict, level, comment, priority, priority_note,
                 n_red, n_yellow, n_d_red, n_d_yellow) = write_excel_bytes(
                    df, agg, counts, severities, delivery, channel_summary, offline_presence,
                    recon, checklist, uploaded.name, months_str, brands_str, len(df))
                status.write(f"Report written in {time.perf_counter() - t3:.1f}s.")

                status.update(label=f"Done in {time.perf_counter() - t0:.1f}s.",
                              state="complete", expanded=False)

            st.session_state.inspect_results = {
                "fatal_error": None,
                "filename": build_output_filename(df, ctx),
                "catalog_warning": catalog_warning,
                "verdict": verdict,
                "level": level,
                "comment": comment,
                "priority": priority,
                "priority_note": priority_note,
                "offline_presence": offline_presence,
                "counts": counts,
                "severities": severities,
                "n_rows": len(df),
                "n_flagged": len(agg),
                "n_delivery": len(delivery),
                "excel_bytes": excel_bytes,
            }
        except Exception as exc:
            st.session_state.inspect_results = {"fatal_error": str(exc)}
        st.rerun()

    results = st.session_state.inspect_results
    if not results:
        return

    if st.button("Start over", key="inspect_reset"):
        reset_all()
        st.rerun()

    st.divider()
    st.subheader("Results")

    if results.get("fatal_error"):
        st.error(results["fatal_error"])
        return

    if results["catalog_warning"]:
        st.warning(results["catalog_warning"])

    level = results["level"]
    banner = st.success if level == "green" else (st.error if level == SEV_RED else st.warning)
    banner(f"{results['verdict'].iloc[0]['Value']} — {results['comment']}")

    st.dataframe(results["verdict"], use_container_width=True, hide_index=True)

    st.subheader(PRIORITY_TITLE)
    if len(results["priority"]):
        st.dataframe(_style_priority(results["priority"]), use_container_width=True, hide_index=True)
    st.caption(results["priority_note"])

    chart = _build_rule_chart(results["counts"], results["severities"])
    if chart is not None:
        st.subheader("Flagged rows by rule")
        st.altair_chart(chart, use_container_width=True)

    st.subheader("Offline/Traditional-non TV presence (by month)")
    st.caption("If these channels aren't bought for this brand, a 'Missing' flag here is expected.")
    st.dataframe(_style_offline_presence(results["offline_presence"]),
                use_container_width=True, hide_index=True)

    st.download_button(
        "Download full QA report (.xlsx)",
        data=results["excel_bytes"],
        file_name=results["filename"],
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
    )
