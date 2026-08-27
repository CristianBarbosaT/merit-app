"""Data Caveats Generator — reads delivery files (RROI or LCA source schema, user-picked
— see FORMAT_PROFILES), detects "Null impressions" / "Null cost" caveats at
placement-month granularity, and writes one Data Caveat Log per brand from the corporate
template (keeps its formatting, table and dropdowns).

Ported from the standalone script `generar_data_caveats.py` (folder-based, run by hand)
into an in-app, upload/download flow — see estado_actual_app.md for the full port notes.
All file I/O happens in memory (BytesIO); nothing is written to disk.
"""
import os
import re
import zipfile
from copy import copy
from datetime import datetime
from io import BytesIO

import openpyxl
import pandas as pd
import streamlit as st

# ============================================================================
# Fixed business rules (Unilever / Mindshare Data Caveat Log workflow)
# ============================================================================
MONTHS_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# Columns of the Data Caveat Log (order A..K of the template's table)
LOG_COLUMNS = ["Type", "Retailer", "Brand", "Category", "Channel", "Campaign",
               "Site", "Placement", "Month", "Description", "Status"]

GROUP_KEYS = ["Brand", "Category", "Channel", "Campaign", "Site", "Placement", "Retailer", "Period"]

DETAIL_COLUMNS = ["Brand", "Category", "Channel", "Campaign", "Site", "Placement",
                  "Month", "Days", "DaysNullCost", "DaysNullImpr",
                  "Impressions", "Cost", "GRPs", "VideoViews"]

# Maps a canonical field name to the column expected in each uploaded delivery file.
# Two source schemas are supported: RROI (the raw-file schema the backfill tool already
# works with) and LCA (the same fields, plus extra columns the caveat detection doesn't
# need, and two columns spelled with a space instead of an underscore).
COLS_RROI = {
    "channel": "Channel", "date": "Date", "brand": "Brand", "category": "Category",
    "campaign": "Prisma_Campaign_Secondary", "site": "Raw_Partner",
    "placement": "Package_Placement_Name", "retailer": "Retailer",
    "impressions": "Impressions", "cost": "Media_Cost",
    # Optional: only used for validation checks; ignored if absent.
    "grps": "GRPs", "video_views": "Video_Views",
}
COLS_LCA = {**COLS_RROI, "cost": "Media Cost", "video_views": "Video Views"}
OPTIONAL_COLS = ("retailer", "grps", "video_views")

FORMAT_PROFILES = {
    "RROI": {"label": "RROI", "cols": COLS_RROI},
    "LCA": {"label": "LCA", "cols": COLS_LCA},
}
DEFAULT_FORMAT = "RROI"


def required_cols(cols_map: dict) -> set:
    return {cols_map[k] for k in cols_map if k not in OPTIONAL_COLS}

# The Month column is written as a real date and displayed in this Excel format, so the
# log shows "Jun-26" rather than a bare "Jun" that loses the year.
MONTH_CELL_FORMAT = "mmm-yy"

TYPE_VALUE = "Null value check"
STATUS_VALUE = "Leave as is - data is correct"
DESC_NULL_IMPRESSIONS = "Null impressions"
DESC_NULL_COST = "Null cost"
BLANK_PLACEHOLDER = "(blank)"
RETAILER_IGNORE = {"(all)", "(blank)", "all", "na", "n/a", ""}
MIN_IMPRESSIONS = 0
MIN_COST = 0

# Cosmetic only — renames a brand for the output file name / tab; never affects detection.
BRAND_NAME_OVERRIDES = {"TRESemme": "Tresemme"}

DETECTION_MODE_LABELS = {
    "month": "Month total (recommended)",
    "row": "Any single day (broader — also catches days a complete month total hides)",
}
DETECTION_MODE_HELP = (
    "**Month total**: a line is a caveat if that placement's WHOLE MONTH total has cost "
    "but no impressions (or the reverse). This is the mode validated against the template.\n\n"
    "**Any single day**: also flags a line if ANY single day matches the pattern, even "
    "though the month total is complete. Finds a lot more (especially in TV and Digital "
    "Social), but includes lines whose monthly total is actually fine."
)

DEFAULT_TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "data_caveats_template.xlsx")

STATE_DEFAULTS = {
    "caveats_data": None,           # concatenated row-level DataFrame across all files
    "caveats_file_summaries": [],   # per-file read results, for the "files read" table
    "caveats_warnings": [],         # non-fatal issues found while reading
    "caveats_errors": [],           # files that failed to read
    "caveats_results": None,        # dict produced by the "Generate" step
}


def init_state():
    for key, value in STATE_DEFAULTS.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_all():
    for key, value in STATE_DEFAULTS.items():
        st.session_state[key] = value


# ----------------------------------------------------------------------------
# Small utilities
# ----------------------------------------------------------------------------
def month_label(year: int, month: int) -> str:
    """(2026, 3) -> "Mar'26" """
    return f"{MONTHS_ABBR[month - 1]}'{year % 100:02d}"


def period_label(period: pd.Period) -> str:
    return month_label(period.year, period.month)


def range_label(start: pd.Period, end: pd.Period) -> str:
    s, e = period_label(start), period_label(end)
    return s if s == e else f"{s}-{e}"


def clean_text(series: pd.Series) -> pd.Series:
    """Normalizes a text column: nulls and blanks become '(blank)'."""
    out = series.astype("object").where(series.notna(), None)
    return out.map(lambda v: BLANK_PLACEHOLDER if v is None or str(v).strip() == ""
                   else str(v).strip())


def display_brand(brand: str) -> str:
    return BRAND_NAME_OVERRIDES.get(brand, brand)


def safe_filename(name: str) -> str:
    return re.sub(r'[<>:"/\\|?*]', "-", str(name)).strip()


def build_sheet_name(inc: str, category: str, brand: str) -> str:
    """Builds the output tab name, respecting Excel's 31-character limit.

    If it doesn't fit, first drops the category, then trims the brand as a last resort
    (never trims the INC#, since that's what identifies the ticket).
    """
    def clean(value):
        return re.sub(r"[\[\]:*?/\\]", "-", str(value)).replace("﻿", "").strip()

    inc, brand = clean(inc), clean(brand)
    cat = clean(category).replace(" ", "")

    name = clean(f"{inc}_{cat}_{brand}")
    if len(name) <= 31:
        return name or "Data Caveats"

    short = clean(f"{inc}__{brand}")
    short = re.sub(r"__+", "_", short).strip("_ ")
    if len(short) <= 31:
        return short

    overflow = len(short) - 31
    trimmed = brand[:max(1, len(brand) - overflow)].rstrip(" _-+")
    short = re.sub(r"__+", "_", clean(f"{inc}__{trimmed}")).strip("_ ")
    return short[:31].rstrip(" _-") or "Data Caveats"


# ----------------------------------------------------------------------------
# Reading uploaded delivery files
# ----------------------------------------------------------------------------
def pick_sheet(file_bytes: bytes, cols_map: dict) -> tuple:
    """Picks the data sheet: the last one (right-most tab) whose header contains every
    required column. Real delivery files consistently put the working data on the last
    tab (a blank/staging "Sheet1" often comes first), so this covers them without extra
    configuration."""
    required = required_cols(cols_map)
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
    try:
        candidates = []
        for ws in wb.worksheets:
            header_row = next(ws.iter_rows(min_row=1, max_row=1), [])
            header = {c.value for c in header_row if c.value is not None}
            if required <= header:
                candidates.append(ws.title)
    finally:
        wb.close()

    if not candidates:
        raise ValueError(
            f"No sheet has all the expected columns ({', '.join(sorted(required))})."
        )
    return candidates[-1], candidates


def read_delivery_file(file_bytes: bytes, filename: str, format_key: str = DEFAULT_FORMAT):
    """Reads one uploaded delivery file in the given source format (RROI or LCA — see
    FORMAT_PROFILES). Returns (row-level DataFrame, sheet used, candidate sheets, rows
    dropped for having no valid date)."""
    cols = FORMAT_PROFILES[format_key]["cols"]
    sheet, candidates = pick_sheet(file_bytes, cols)
    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet, engine="openpyxl")

    missing = [cols[k] for k in cols if k not in OPTIONAL_COLS and cols[k] not in df.columns]
    if missing:
        raise ValueError(f"Missing columns in sheet '{sheet}': {missing}")

    out = pd.DataFrame()
    out["Brand"] = clean_text(df[cols["brand"]])
    out["Category"] = clean_text(df[cols["category"]])
    out["Channel"] = clean_text(df[cols["channel"]])
    out["Campaign"] = clean_text(df[cols["campaign"]])
    out["Site"] = clean_text(df[cols["site"]])
    out["Placement"] = clean_text(df[cols["placement"]])
    out["Retailer"] = (clean_text(df[cols["retailer"]]) if cols["retailer"] in df.columns
                       else BLANK_PLACEHOLDER)

    dates = pd.to_datetime(df[cols["date"]], errors="coerce")
    out["Period"] = dates.dt.to_period("M")

    def numeric(key):
        col = cols.get(key)
        if col and col in df.columns:
            return pd.to_numeric(df[col], errors="coerce").fillna(0)
        return pd.Series(0.0, index=df.index)

    out["Impressions"] = numeric("impressions")
    out["Cost"] = numeric("cost")
    out["GRPs"] = numeric("grps")
    out["VideoViews"] = numeric("video_views")

    # Row-level (daily) flags — used by validation and by DETECTION_MODE == "row"
    out["RowNullCost"] = ((out["Impressions"] > MIN_IMPRESSIONS) &
                          (out["Cost"] <= MIN_COST)).astype(int)
    out["RowNullImpr"] = ((out["Cost"] > MIN_COST) &
                          (out["Impressions"] <= MIN_IMPRESSIONS)).astype(int)
    out["SourceFile"] = filename

    dropped = int(out["Period"].isna().sum())
    out = out[out["Period"].notna()].copy()
    return out, sheet, candidates, dropped


@st.cache_data(show_spinner=False)
def _read_delivery_file_cached(file_bytes: bytes, filename: str, format_key: str = DEFAULT_FORMAT):
    return read_delivery_file(file_bytes, filename, format_key)


# ----------------------------------------------------------------------------
# Caveat detection
# ----------------------------------------------------------------------------
def build_groups(df: pd.DataFrame, filter_by_range: bool, range_start: pd.Period,
                  range_end: pd.Period) -> pd.DataFrame:
    """Aggregates to placement-month grain and flags the caveat patterns."""
    grouped = df.groupby(GROUP_KEYS, dropna=False, observed=True).agg(
        Impressions=("Impressions", "sum"),
        Cost=("Cost", "sum"),
        GRPs=("GRPs", "sum"),
        VideoViews=("VideoViews", "sum"),
        Days=("Impressions", "size"),
        DaysNullCost=("RowNullCost", "sum"),
        DaysNullImpr=("RowNullImpr", "sum"),
    ).reset_index()

    if filter_by_range:
        grouped = grouped[(grouped["Period"] >= range_start) & (grouped["Period"] <= range_end)].copy()

    has_impressions = grouped["Impressions"] > MIN_IMPRESSIONS
    has_cost = grouped["Cost"] > MIN_COST

    # Pattern at the whole-month total
    grouped["MonthNullImpr"] = has_cost & ~has_impressions
    grouped["MonthNullCost"] = has_impressions & ~has_cost
    # Individual days matching the pattern whose month total is nonetheless complete
    grouped["MaskedNullCost"] = (grouped["DaysNullCost"] > 0) & ~grouped["MonthNullCost"]
    grouped["MaskedNullImpr"] = (grouped["DaysNullImpr"] > 0) & ~grouped["MonthNullImpr"]
    # Lines at zero total that nonetheless show activity
    grouped["ZeroButActive"] = (~has_impressions & ~has_cost &
                                ((grouped["GRPs"] > 0) | (grouped["VideoViews"] > 0)))
    return grouped.reset_index(drop=True)


def classify(grouped: pd.DataFrame, detection_mode: str) -> pd.DataFrame:
    """Assigns the caveat description according to detection_mode."""
    grouped = grouped.copy()
    grouped["Description"] = pd.NA
    grouped.loc[grouped["MonthNullImpr"], "Description"] = DESC_NULL_IMPRESSIONS
    grouped.loc[grouped["MonthNullCost"], "Description"] = DESC_NULL_COST

    if detection_mode == "row":
        pending = grouped["Description"].isna()
        grouped.loc[pending & (grouped["DaysNullCost"] >= grouped["DaysNullImpr"]) &
                    (grouped["DaysNullCost"] > 0), "Description"] = DESC_NULL_COST
        grouped.loc[grouped["Description"].isna() & (grouped["DaysNullImpr"] > 0),
                    "Description"] = DESC_NULL_IMPRESSIONS

    caveats = grouped[grouped["Description"].notna()].copy()
    caveats["Type"] = TYPE_VALUE
    caveats["Status"] = STATUS_VALUE
    # A real date (the month's first day), not a "Jun" string: the template's own Month
    # dropdown (column I, source $Y$3:$Y$13) holds datetimes, and an abbreviation alone
    # loses the year — two Junes from different years were indistinguishable in the log.
    # write_brand_file_bytes formats the cell as "mmm-yy" so it still reads as "Jun-26".
    caveats["Month"] = caveats["Period"].dt.to_timestamp()
    caveats["Retailer"] = caveats["Retailer"].map(
        lambda v: "" if str(v).strip().lower() in RETAILER_IGNORE else v)

    caveats["_desc_order"] = (caveats["Description"] == DESC_NULL_COST).astype(int)
    caveats = caveats.sort_values(
        ["_desc_order", "Channel", "Campaign", "Site", "Placement", "Period"]
    ).reset_index(drop=True)

    return caveats[LOG_COLUMNS]


# ----------------------------------------------------------------------------
# Validation
# ----------------------------------------------------------------------------
def _detail(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    out["Month"] = out["Period"].astype(str)
    return out[DETAIL_COLUMNS]


def validate(data: pd.DataFrame, grouped: pd.DataFrame, caveats: pd.DataFrame, brands: list,
             filter_by_range: bool):
    """Checks the data before writing any files. Returns (issues, detail) where issues is
    a list of {level, key, title, count, note} and detail a dict {sheet_name: DataFrame}."""
    issues, detail = [], {}

    def add(level, key, title, count, note, frame=None):
        issues.append({"level": level, "key": key, "title": title, "count": count, "note": note})
        if frame is not None and len(frame):
            detail[key] = _detail(frame)

    # 1) Channels with no Campaign/Placement -> the detection grain there is coarser
    coarse = (data.groupby("Channel", observed=True)
              .agg(rows=("Placement", "size"),
                   no_placement=("Placement", lambda s: (s == BLANK_PLACEHOLDER).sum()))
              .reset_index())
    coarse = coarse[coarse["no_placement"] == coarse["rows"]]
    if len(coarse):
        channels = ", ".join(f"{r.Channel} ({r.rows:,} rows)" for r in coarse.itertuples())
        add("INFO", "coarse_grain",
            "Channels with no Campaign/Placement in the input", len(coarse),
            f"The finest possible line there is Partner + Month: {channels}")

    # 2) Caveats masked by the monthly aggregation
    masked = grouped[grouped["MaskedNullCost"] | grouped["MaskedNullImpr"]]
    if len(masked):
        days = int(masked["DaysNullCost"].sum() + masked["DaysNullImpr"].sum())
        add("WARNING", "masked",
            "Lines with individual days missing cost/impressions that the month total hides",
            len(masked),
            f"{days:,} individual days across {len(masked)} lines. Their month total DOES "
            "have both cost and impressions, so they don't count as a caveat under "
            "detection mode 'Month total'.", masked)

    # 3) Lines at zero total but with activity (GRPs / Video Views)
    zero_active = grouped[grouped["ZeroButActive"]]
    if len(zero_active):
        add("WARNING", "zero_but_active",
            "Lines with GRPs/Video Views but both impressions AND cost at zero",
            len(zero_active),
            "Don't count as a caveat because they're missing both metrics at once. "
            "Worth checking whether they should be reported.", zero_active)

    # 4) Negative values
    negatives = grouped[(grouped["Impressions"] < 0) | (grouped["Cost"] < 0)]
    if len(negatives):
        add("WARNING", "negatives", "Lines with negative impressions or cost",
            len(negatives), "Usually adjustments or credits — worth a check.", negatives)

    # 5) Brands with no caveat lines at all
    no_caveats = [b for b in brands if not len(caveats[caveats["Brand"] == b])]
    if no_caveats:
        add("FLAG", "brands_no_caveats", "Brands with no caveat lines at all",
            len(no_caveats),
            "Double-check manually that the delivery for these brands is complete: "
            + ", ".join(display_brand(b) for b in no_caveats))

    # 6) The selected range leaves nothing
    if filter_by_range and not len(grouped[grouped["Days"] > 0]):
        add("ERROR", "empty_range", "The selected date range leaves no rows at all",
            0, "Check the date range against the dates actually in the input.")

    return issues, detail


def write_validation_report_bytes(issues: list, detail: dict) -> bytes:
    summary = pd.DataFrame([
        {"Level": i["level"], "Finding": i["title"], "Lines": i["count"], "Detail": i["note"]}
        for i in issues
    ])
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        (summary if len(summary) else pd.DataFrame(
            [{"Level": "OK", "Finding": "No findings", "Lines": 0, "Detail": ""}]
        )).to_excel(writer, sheet_name="Summary", index=False)
        for key, frame in detail.items():
            frame.to_excel(writer, sheet_name=key[:31], index=False)
    return output.getvalue()


# ----------------------------------------------------------------------------
# Writing the output file (from the template)
# ----------------------------------------------------------------------------
def find_log_table(ws):
    """Returns (table, header_row, first_data_row, last_data_row)."""
    tables = list(ws.tables.values()) if hasattr(ws, "tables") else []
    if tables:
        table = tables[0]
        first_cell, last_cell = table.ref.split(":")
        header_row = int(re.sub(r"[A-Z]", "", first_cell))
        last_row = int(re.sub(r"[A-Z]", "", last_cell))
        return table, header_row, header_row + 1, last_row

    for row in ws.iter_rows(min_row=1, max_row=30, max_col=1):
        if str(row[0].value).strip() == "Type":
            header_row = row[0].row
            return None, header_row, header_row + 1, ws.max_row
    raise ValueError("Could not locate the Data Caveat Log table in the template.")


def write_brand_file_bytes(caveats: pd.DataFrame, brand: str, category: str,
                            template_bytes: bytes, inc_number: str) -> tuple:
    """Returns (xlsx bytes, tab name actually used)."""
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    ws = wb.worksheets[0]

    table, header_row, first_data_row, template_last_row = find_log_table(ws)
    n_cols = 14  # A..N (includes the template's Column1/2/3)

    row_styles = [copy(ws.cell(row=first_data_row, column=c)._style) for c in range(1, n_cols + 1)]
    blank_style = copy(ws.cell(row=template_last_row + 3, column=1)._style)

    for r in range(first_data_row, template_last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).value = None

    for i, (_, rec) in enumerate(caveats.iterrows()):
        r = first_data_row + i
        values = [rec["Type"], rec["Retailer"], rec["Brand"], rec["Category"], rec["Channel"],
                  rec["Campaign"], rec["Site"], rec["Placement"], rec["Month"],
                  rec["Description"], rec["Status"], None, None, None]
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c)
            cell.value = value
            cell._style = copy(row_styles[c - 1])
            if c == 9:  # Month: a real date, displayed as "Jun-26"
                cell.number_format = MONTH_CELL_FORMAT

    n_rows = len(caveats)
    last_data_row = first_data_row + max(n_rows, 1) - 1  # the table needs >= 1 row

    if n_rows == 0:  # leave one blank but formatted row
        for c in range(1, n_cols + 1):
            ws.cell(row=first_data_row, column=c)._style = copy(row_styles[c - 1])

    for r in range(last_data_row + 1, template_last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c)._style = copy(blank_style)

    if table is not None:
        table.ref = f"A{header_row}:N{last_data_row}"
    for dv in ws.data_validations.dataValidation:
        ranges = str(dv.sqref)
        col = ranges[0] if ranges else None
        if col in ("A", "I"):
            dv.sqref = f"{col}{first_data_row}:{col}{last_data_row}"

    ws.title = build_sheet_name(inc_number, category, display_brand(brand))

    output = BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue(), ws.title


# ----------------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------------
def render():
    if st.button("← Back to menu"):
        st.session_state.active_tool = None
        st.rerun()

    init_state()

    st.title("Data Caveats Generator")
    st.caption(
        "Reads your delivery files, detects rows with cost but no impressions (or the "
        "reverse) at the placement-month level, and produces one Data Caveat Log per "
        "brand ready to send — built from the corporate template, formatting and all."
    )

    format_key = st.radio(
        "Delivery file format",
        list(FORMAT_PROFILES),
        format_func=lambda k: FORMAT_PROFILES[k]["label"],
        horizontal=True, key="caveats_format",
    )

    uploaded_files = st.file_uploader(
        "Upload delivery files (.xlsx) — one or more, one per brand or combined",
        type=["xlsx"], accept_multiple_files=True, key="caveats_uploader",
    )

    with st.expander("Advanced: use a different template"):
        st.caption(
            "Leave this empty to use the built-in Data Caveat Log template. Only override "
            "it if you have a different approved version — the table and dropdowns must "
            "match what this tool expects (see estado_actual_app.md)."
        )
        custom_template = st.file_uploader(
            "Template (.xlsx)", type=["xlsx"], key="caveats_template_uploader"
        )

    if uploaded_files and st.button("Read files", type="primary"):
        frames, summaries, warnings, errors = [], [], [], []
        for uploaded in uploaded_files:
            file_bytes = uploaded.getvalue()
            try:
                df, sheet, candidates, dropped = _read_delivery_file_cached(
                    file_bytes, uploaded.name, format_key
                )
            except Exception as exc:
                errors.append(f"{uploaded.name}: {exc}")
                continue

            brands = sorted(df["Brand"].unique())
            summaries.append({
                "File": uploaded.name, "Format": format_key, "Sheet": sheet, "Rows": len(df),
                "Brand(s)": ", ".join(brands),
            })
            if len(candidates) > 1:
                warnings.append(
                    f"{uploaded.name}: several candidate sheets {candidates} -> used '{sheet}'"
                )
            if dropped:
                warnings.append(f"{uploaded.name}: {dropped} rows with no valid date were ignored")
            frames.append(df)

        if not frames:
            st.error("None of the uploaded files could be read. See the errors below.")
            for e in errors:
                st.error(e)
        else:
            st.session_state.caveats_data = pd.concat(frames, ignore_index=True)
            st.session_state.caveats_file_summaries = summaries
            st.session_state.caveats_warnings = warnings
            st.session_state.caveats_errors = errors
            st.session_state.caveats_results = None
        st.rerun()

    data = st.session_state.caveats_data
    if data is None:
        return

    if st.button("Start over"):
        reset_all()
        st.rerun()

    st.divider()
    st.subheader(f"Files read ({len(st.session_state.caveats_file_summaries)})")
    st.dataframe(
        pd.DataFrame(st.session_state.caveats_file_summaries),
        use_container_width=True, hide_index=True,
    )
    for w in st.session_state.caveats_warnings:
        st.warning(w)
    for e in st.session_state.caveats_errors:
        st.error(e)

    st.divider()
    st.subheader("Settings")

    periods = sorted(data["Period"].dropna().unique())
    if not periods:
        st.error("None of the uploaded rows have a valid date.")
        return
    labels = [period_label(p) for p in periods]
    label_to_period = dict(zip(labels, periods))

    if len(labels) > 1:
        start_label, end_label = st.select_slider(
            "Date range to report", options=labels, value=(labels[0], labels[-1]),
        )
    else:
        start_label = end_label = labels[0]
        st.caption(f"Only one month in the uploaded data: {labels[0]}.")
    range_start, range_end = label_to_period[start_label], label_to_period[end_label]
    rng_label = range_label(range_start, range_end)

    detection_mode = st.radio(
        "Detection granularity", list(DETECTION_MODE_LABELS),
        format_func=lambda k: DETECTION_MODE_LABELS[k], key="caveats_detection_mode",
    )
    st.caption(DETECTION_MODE_HELP)

    inc_number = st.text_input(
        "INC# for the output tabs (leave blank to keep the 'INC#' placeholder)",
        key="caveats_inc_number",
    ).strip() or "INC#"

    opt_col1, opt_col2 = st.columns(2)
    with opt_col1:
        filter_by_range = st.checkbox(
            "Filter the data to the selected range", value=True, key="caveats_filter_by_range",
        )
        run_validation = st.checkbox(
            "Run validation before generating", value=True, key="caveats_run_validation",
        )
    with opt_col2:
        generate_if_no_caveats = st.checkbox(
            "Generate a file even if a brand has no caveats (flagged either way)",
            value=True, key="caveats_generate_if_no_caveats",
        )
        stop_on_issues = st.checkbox(
            "Stop instead of generating if validation finds issues",
            value=False, key="caveats_stop_on_issues",
        )

    if st.button("Generate Data Caveat Logs", type="primary", use_container_width=True):
        template_bytes = (
            custom_template.getvalue() if custom_template is not None
            else open(DEFAULT_TEMPLATE_PATH, "rb").read()
        )

        grouped = build_groups(data, filter_by_range, range_start, range_end)
        caveats = classify(grouped, detection_mode)
        all_brands = sorted(data["Brand"].unique())

        issues, detail = [], {}
        if run_validation:
            issues, detail = validate(data, grouped, caveats, all_brands, filter_by_range)
            blocking = [i for i in issues if i["level"] in ("ERROR", "WARNING")]
            if blocking and stop_on_issues:
                st.session_state.caveats_results = {
                    "stopped": True, "issues": issues, "range_label": rng_label,
                }
                st.rerun()

        summary_rows, flags, file_warnings, file_errors, output_files = [], [], [], [], {}
        for brand in all_brands:
            brand_rows = caveats[caveats["Brand"] == brand].copy()
            categories = sorted(data.loc[data["Brand"] == brand, "Category"].unique())
            if len(categories) > 1:
                category = " & ".join(categories)
                file_warnings.append(
                    f"{brand}: has several categories {categories} -- kept as-is, each row "
                    "keeps its own category; not standardized to a single one"
                )
            else:
                category = categories[0] if categories else ""

            n_imp = int((brand_rows["Description"] == DESC_NULL_IMPRESSIONS).sum())
            n_cost = int((brand_rows["Description"] == DESC_NULL_COST).sum())
            total = len(brand_rows)

            filename = safe_filename(f"{display_brand(brand)}_DataCaveatLog_{rng_label}.xlsx")

            if total == 0:
                flags.append(brand)
                if not generate_if_no_caveats:
                    summary_rows.append({
                        "Brand": display_brand(brand), "Category": category, "Caveats": 0,
                        "Null Impr": 0, "Null Cost": 0, "Status": "NO CAVEATS - not generated",
                    })
                    continue

            try:
                file_bytes, sheet_name = write_brand_file_bytes(
                    brand_rows, brand, category, template_bytes, inc_number
                )
            except Exception as exc:
                file_errors.append(f"{brand}: could not generate the file -> {exc}")
                summary_rows.append({
                    "Brand": display_brand(brand), "Category": category, "Caveats": total,
                    "Null Impr": n_imp, "Null Cost": n_cost, "Status": "ERROR",
                })
                continue

            output_files[filename] = file_bytes
            status = "OK" if total else "OK (empty log - FLAG)"
            summary_rows.append({
                "Brand": display_brand(brand), "Category": category, "Caveats": total,
                "Null Impr": n_imp, "Null Cost": n_cost, "Status": status,
            })

        validation_bytes = write_validation_report_bytes(issues, detail) if run_validation else None

        summary_lines = [
            f"Data Caveat Logs summary - {rng_label}",
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Detection granularity: {detection_mode}", "",
        ]
        for row in summary_rows:
            summary_lines.append(
                f"{row['Brand']:<18} {str(row['Category'])[:12]:<12} {row['Caveats']:>4} "
                f"caveats  (null impr: {row['Null Impr']}, null cost: {row['Null Cost']})  "
                f"{row['Status']}"
            )
        if flags:
            summary_lines.append("")
            summary_lines.append("FLAG - brands with no caveat lines: " +
                                 ", ".join(display_brand(b) for b in flags))
        summary_txt = "\n".join(summary_lines).encode("utf-8")

        zip_buffer = BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
            for filename, file_bytes in output_files.items():
                zf.writestr(filename, file_bytes)
            if validation_bytes:
                zf.writestr(f"_Validation_DataCaveats_{safe_filename(rng_label)}.xlsx", validation_bytes)
            zf.writestr(f"_Summary_DataCaveats_{safe_filename(rng_label)}.txt", summary_txt)

        st.session_state.caveats_results = {
            "stopped": False,
            "issues": issues,
            "summary_rows": summary_rows,
            "flags": flags,
            "warnings": file_warnings,
            "errors": file_errors,
            "zip_bytes": zip_buffer.getvalue(),
            "range_label": rng_label,
            "n_files": len(output_files),
        }
        st.rerun()

    results = st.session_state.caveats_results
    if not results:
        return

    st.divider()
    st.subheader("Results")

    if results.get("stopped"):
        st.error(
            "Generation stopped: validation found issues and 'Stop instead of generating' "
            "is on. Review the findings below, fix the input if needed, and either turn "
            "that option off or re-upload the corrected files."
        )

    if results["issues"]:
        level_order = {"ERROR": 0, "FLAG": 1, "WARNING": 2, "INFO": 3}
        level_fn = {"ERROR": st.error, "FLAG": st.warning, "WARNING": st.warning, "INFO": st.info}
        for issue in sorted(results["issues"], key=lambda i: level_order.get(i["level"], 9)):
            level_fn.get(issue["level"], st.info)(
                f"**[{issue['level']}] {issue['title']}** ({issue['count']})\n\n{issue['note']}"
            )
    elif not results.get("stopped"):
        st.success("Validation found no issues. The data looks consistent.")

    if results.get("stopped"):
        return

    st.subheader(f"Summary — {results['range_label']}")
    st.dataframe(pd.DataFrame(results["summary_rows"]), use_container_width=True, hide_index=True)

    if results["flags"]:
        st.warning(
            "Brands with no caveat lines at all — double-check their delivery is complete: "
            + ", ".join(display_brand(b) for b in results["flags"])
        )
    for w in results["warnings"]:
        st.warning(w)
    for e in results["errors"]:
        st.error(e)

    st.download_button(
        f"Download all files (.zip) — {results['n_files']} brand file(s)",
        data=results["zip_bytes"],
        file_name=f"DataCaveats_{safe_filename(results['range_label'])}.zip",
        mime="application/zip",
        type="primary",
    )
