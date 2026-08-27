"""Tests for the Data Caveats Generator (tools/data_caveats.py), ported from the
standalone generar_data_caveats.py script. Builds small synthetic delivery files
in-memory (matching the real RROI file schema) and exercises the pure pipeline
functions directly, plus one pass against the real bundled template."""
from io import BytesIO

import openpyxl
import pandas as pd

from tools.data_caveats import (
    DEFAULT_TEMPLATE_PATH, DESC_NULL_COST, DESC_NULL_IMPRESSIONS, BLANK_PLACEHOLDER,
    MONTH_CELL_FORMAT, STATUS_VALUE, TYPE_VALUE, build_groups, classify, display_brand,
    month_label, period_label, range_label, read_delivery_file, validate,
    write_brand_file_bytes, build_sheet_name, safe_filename,
)

SOURCE_COLUMNS = ["Channel", "Date", "Brand", "Category", "Prisma_Campaign_Secondary",
                  "Raw_Partner", "Package_Placement_Name", "Retailer", "Impressions",
                  "Media_Cost", "GRPs", "Video_Views"]


def row(date, brand="BrandA", category="Hair Care", channel="Digital Social",
        campaign="CampX", site="PINTEREST", placement="PL1_UNE", retailer="(all)",
        impressions=None, media_cost=None, grps=None, video_views=None):
    return {
        "Channel": channel, "Date": date, "Brand": brand, "Category": category,
        "Prisma_Campaign_Secondary": campaign, "Raw_Partner": site,
        "Package_Placement_Name": placement, "Retailer": retailer,
        "Impressions": impressions, "Media_Cost": media_cost, "GRPs": grps,
        "Video_Views": video_views,
    }


def make_upload_bytes(rows):
    buf = BytesIO()
    pd.DataFrame(rows).to_excel(buf, index=False)
    return buf.getvalue()


JUNE = pd.Period("2026-06", freq="M")
MAY = pd.Period("2026-05", freq="M")

# ---------------------------------------------------------------------------
# TEST A: reading a delivery file — parsing, cleaning, numeric coercion
# ---------------------------------------------------------------------------
raw_rows = [
    row("2026-06-01", impressions=1000, media_cost=None),
    row("2026-06-02", impressions=None, media_cost=None, campaign=None),  # -> "(blank)"
    row("2026-06-03", impressions="not a date placeholder"[:0] or None),  # blank text -> 0
    {**row("2026-06-04"), "Date": "not-a-date"},  # invalid date -> dropped
]
df, sheet, candidates, dropped = read_delivery_file(make_upload_bytes(raw_rows), "test.xlsx")
assert sheet == "Sheet1" and candidates == ["Sheet1"]
assert dropped == 1, f"expected exactly 1 row dropped for an invalid date, got {dropped}"
assert len(df) == 3
assert df["Campaign"].iloc[1] == BLANK_PLACEHOLDER, "blank Campaign must become '(blank)'"
assert df["Impressions"].fillna(0).tolist()[0] == 1000.0
assert df["Impressions"].dtype.kind == "f"
print("TEST A OK: read_delivery_file parses, cleans and coerces correctly")

# ---------------------------------------------------------------------------
# TEST B: month-level caveat detection — the two core patterns
# ---------------------------------------------------------------------------
basic_rows = [
    # PL1: cost present, impressions never -> "Null impressions" for the month
    row("2026-06-01", placement="PL1_UNE", impressions=None, media_cost=5.0),
    row("2026-06-02", placement="PL1_UNE", impressions=None, media_cost=7.0),
    # PL2: impressions present, cost never -> "Null cost" for the month
    row("2026-06-01", placement="PL2_UNE", impressions=100, media_cost=None),
    row("2026-06-02", placement="PL2_UNE", impressions=200, media_cost=None),
    # PL3: both present every day -> not a caveat
    row("2026-06-01", placement="PL3_UNE", impressions=50, media_cost=2.0),
    row("2026-06-02", placement="PL3_UNE", impressions=60, media_cost=3.0),
]
df_b, *_ = read_delivery_file(make_upload_bytes(basic_rows), "b.xlsx")
grouped_b = build_groups(df_b, filter_by_range=True, range_start=JUNE, range_end=JUNE)
caveats_b = classify(grouped_b)

by_placement = dict(zip(caveats_b["Placement"], caveats_b["Description"]))
assert by_placement.get("PL1_UNE") == DESC_NULL_IMPRESSIONS
assert by_placement.get("PL2_UNE") == DESC_NULL_COST
assert "PL3_UNE" not in by_placement, "a placement with both metrics present is not a caveat"
assert (caveats_b["Type"] == TYPE_VALUE).all()
assert (caveats_b["Status"] == STATUS_VALUE).all()
# Month is a real date (the month's first day), so the year survives — a bare "Jun"
# could not tell Jun'25 from Jun'26 in a log spanning more than one year
assert set(caveats_b["Month"]) == {pd.Timestamp("2026-06-01")}, set(caveats_b["Month"])
print("TEST B OK: 'Null impressions' / 'Null cost' detected correctly at month grain, "
      "a complete placement is not flagged")

# ---------------------------------------------------------------------------
# TEST C: a day-level gap that the month total hides is NOT a caveat -- detection
# is month-total only, with no day-level mode to catch it instead. The day-level
# flags survive (DaysNullImpr, MaskedNullImpr) since they still feed the "masked"
# validation finding (an FYI, not a generated caveat) -- see TEST D.
# ---------------------------------------------------------------------------
masked_rows = [
    row("2026-06-01", placement="PL4_UNE", impressions=100, media_cost=5.0),
    row("2026-06-02", placement="PL4_UNE", impressions=0, media_cost=5.0),  # this day has no impressions
]
df_c, *_ = read_delivery_file(make_upload_bytes(masked_rows), "c.xlsx")
grouped_c = build_groups(df_c, filter_by_range=True, range_start=JUNE, range_end=JUNE)

assert grouped_c["MonthNullImpr"].iloc[0] == False  # noqa: E712 -- month total has impressions
assert grouped_c["DaysNullImpr"].iloc[0] == 1
assert grouped_c["MaskedNullImpr"].iloc[0] == True  # noqa: E712

caveats_month = classify(grouped_c)
assert len(caveats_month) == 0, "the month total is complete, so it must not be flagged"
print("TEST C OK: a masked single-day gap never becomes a caveat -- only the month total counts")

# ---------------------------------------------------------------------------
# TEST D: validate() — each finding in isolation
# ---------------------------------------------------------------------------
validate_rows = [
    # coarse grain: an entire channel with no Campaign/Placement at all
    row("2026-06-01", channel="TV", campaign=None, placement=None, impressions=500, media_cost=20.0),
    row("2026-06-02", channel="TV", campaign=None, placement=None, impressions=600, media_cost=25.0),
    # zero total but real activity (GRPs) -> should NOT be a caveat, should be flagged
    row("2026-06-01", placement="PL5_UNE", impressions=0, media_cost=0.0, grps=5.0),
    # negative values -> flagged, and does not coincidentally match a caveat pattern
    # (both impressions and cost are <=0, so neither MonthNullImpr nor MonthNullCost fires)
    row("2026-06-01", placement="PL6_UNE", impressions=-10, media_cost=0.0),
    # a brand with real, complete data and zero caveats
    row("2026-06-01", brand="BrandB", placement="PLB1_UNE", impressions=40, media_cost=1.0),
    row("2026-06-02", brand="BrandB", placement="PLB1_UNE", impressions=45, media_cost=1.2),
    # BrandA needs at least one real caveat too, so BrandA does NOT show up in the
    # "brands with no caveats" flag (proving that check is brand-specific)
    row("2026-06-01", placement="PL7_UNE", impressions=None, media_cost=9.0),
]
df_d, *_ = read_delivery_file(make_upload_bytes(validate_rows), "d.xlsx")
grouped_d = build_groups(df_d, filter_by_range=True, range_start=JUNE, range_end=JUNE)
caveats_d = classify(grouped_d)
brands_d = sorted(df_d["Brand"].unique())
issues, detail = validate(df_d, grouped_d, caveats_d, brands_d, filter_by_range=True)

by_key = {i["key"]: i for i in issues}
assert "coarse_grain" in by_key and by_key["coarse_grain"]["level"] == "INFO"
assert "TV" in by_key["coarse_grain"]["note"]

assert "zero_but_active" in by_key and by_key["zero_but_active"]["count"] == 1
assert detail["zero_but_active"]["Placement"].iloc[0] == "PL5_UNE"

assert "negatives" in by_key and by_key["negatives"]["count"] == 1
assert detail["negatives"]["Placement"].iloc[0] == "PL6_UNE"
assert "PL6_UNE" not in dict(zip(caveats_d["Placement"], caveats_d["Description"])), (
    "an all-non-positive row must not also register as a Null-impressions caveat"
)

assert "brands_no_caveats" in by_key
assert "BrandB" in by_key["brands_no_caveats"]["note"]
assert "BrandA" not in by_key["brands_no_caveats"]["note"]

assert "masked" not in by_key, "no masked-day scenario was set up in this dataset"
assert "empty_range" not in by_key, "the range is not empty here"
print("TEST D OK: validate() reports coarse-grain channels, zero-but-active, negatives "
      "and no-caveat brands independently, each scoped correctly")

# ---------------------------------------------------------------------------
# TEST E: filter_by_range actually filters
# ---------------------------------------------------------------------------
range_rows = [
    row("2026-06-01", placement="PL_JUN_UNE", impressions=100, media_cost=None),  # June caveat
    row("2026-05-01", placement="PL_MAY_UNE", impressions=None, media_cost=50.0),  # May caveat
]
df_e, *_ = read_delivery_file(make_upload_bytes(range_rows), "e.xlsx")

grouped_filtered = build_groups(df_e, filter_by_range=True, range_start=JUNE, range_end=JUNE)
caveats_filtered = classify(grouped_filtered)
assert set(caveats_filtered["Placement"]) == {"PL_JUN_UNE"}, (
    "filtering to June must exclude the May-only placement entirely"
)

grouped_unfiltered = build_groups(df_e, filter_by_range=False, range_start=JUNE, range_end=JUNE)
caveats_unfiltered = classify(grouped_unfiltered)
assert set(caveats_unfiltered["Placement"]) == {"PL_JUN_UNE", "PL_MAY_UNE"}, (
    "without filtering, both months' caveats must be present"
)
print("TEST E OK: filter_by_range includes/excludes months as expected")

# ---------------------------------------------------------------------------
# TEST F: writing the real output file, against the actual bundled template
# ---------------------------------------------------------------------------
with open(DEFAULT_TEMPLATE_PATH, "rb") as f:
    template_bytes = f.read()

file_bytes, tab_name = write_brand_file_bytes(
    caveats_b, "BrandA", "Hair Care", template_bytes, "INC1234567"
)
assert tab_name == build_sheet_name("INC1234567", "Hair Care", "BrandA")

wb = openpyxl.load_workbook(BytesIO(file_bytes))
assert wb.sheetnames == [tab_name]
ws = wb.worksheets[0]

table = list(ws.tables.values())[0]
first_cell, last_cell = table.ref.split(":")
assert first_cell == "A7"
# 2 caveat rows (PL1 + PL2) written starting at row 8
assert last_cell == "N9", f"expected the table to shrink to exactly 2 data rows, got {last_cell}"

written = [
    [ws.cell(row=r, column=c).value for c in range(1, 12)]
    for r in (8, 9)
]
descriptions = {r[7]: r[9] for r in written}  # Placement -> Description
assert descriptions["PL1_UNE"] == DESC_NULL_IMPRESSIONS
assert descriptions["PL2_UNE"] == DESC_NULL_COST
assert all(r[0] == TYPE_VALUE and r[10] == STATUS_VALUE for r in written)
print("TEST F OK: write_brand_file_bytes produces a correctly-shrunk table with the "
      "right rows, against the real bundled template")

# an empty caveat list still produces a valid (empty but formatted) table, not a crash
empty_bytes, _ = write_brand_file_bytes(
    caveats_b.iloc[0:0], "BrandC", "Hair Care", template_bytes, "INC1234567"
)
wb_empty = openpyxl.load_workbook(BytesIO(empty_bytes))
table_empty = list(wb_empty.worksheets[0].tables.values())[0]
assert table_empty.ref == "A7:N8", "an empty brand must still leave exactly 1 formatted row"
print("TEST F2 OK: a brand with zero caveats still produces a valid formatted file")

# ---------------------------------------------------------------------------
# TEST G: small formatting helpers
# ---------------------------------------------------------------------------
assert month_label(2026, 3) == "Mar'26"
assert period_label(pd.Period("2026-03", freq="M")) == "Mar'26"
assert range_label(pd.Period("2026-01", freq="M"), pd.Period("2026-06", freq="M")) == "Jan'26-Jun'26"
assert range_label(JUNE, JUNE) == "Jun'26"
assert display_brand("TRESemme") == "Tresemme"
assert display_brand("Dove") == "Dove"
assert safe_filename('Bad:Name/With*Chars?') == "Bad-Name-With-Chars-"
long_name = build_sheet_name("INC1234567", "Hair Care", "A Very Long Brand Name Indeed")
assert len(long_name) <= 31
print("TEST G OK: label/filename/sheet-name helpers behave as expected")

# ---------------------------------------------------------------------------
# TEST H: the LCA source format — extra columns, and Media Cost / Video Views
# spelled with a space instead of an underscore
# ---------------------------------------------------------------------------
LCA_COLUMNS = ["Channel", "Partnership", "Date", "Brand", "Campaign",
               "Prisma_Campaign_Secondary", "Product_Line", "Category", "Subcategory",
               "Format", "Raw_Partner", "Audience", "Package_Placement_Name", "Daypart",
               "Retailer", "Breakout", "Impressions", "Clicks", "Media Cost",
               "Video Views", "GRPs"]


def lca_row(date, brand="BrandA", category="Hair Care", channel="Digital Video",
            campaign="CampX", site="YOUTUBE.COM", placement="PL1_UUT", retailer="(all)",
            impressions=None, media_cost=None, grps=None, video_views=None):
    return {
        "Channel": channel, "Partnership": None, "Date": date, "Brand": brand,
        "Campaign": "Some Campaign (LCA)", "Prisma_Campaign_Secondary": campaign,
        "Product_Line": "Hair Care All Other_(all)", "Category": category,
        "Subcategory": "Hair Care All Other", "Format": "(all)", "Raw_Partner": site,
        "Audience": "Custom: Competitive Purchasers", "Package_Placement_Name": placement,
        "Daypart": "NULL", "Retailer": retailer, "Breakout": "NULL",
        "Impressions": impressions, "Clicks": "NULL", "Media Cost": media_cost,
        "Video Views": video_views, "GRPs": grps,
    }


lca_rows = [
    lca_row("2026-06-01", placement="PL1_UUT", impressions=None, media_cost=5.0),
    lca_row("2026-06-02", placement="PL1_UUT", impressions=None, media_cost=7.0),
]
lca_buf = BytesIO()
pd.DataFrame(lca_rows, columns=LCA_COLUMNS).to_excel(lca_buf, index=False)

df_lca, sheet_lca, _, dropped_lca = read_delivery_file(
    lca_buf.getvalue(), "lca_test.xlsx", format_key="LCA"
)
assert dropped_lca == 0
assert len(df_lca) == 2
assert df_lca["Cost"].tolist() == [5.0, 7.0], "Media Cost (space) must map to Cost under LCA"
assert df_lca["Impressions"].tolist() == [0.0, 0.0]

try:
    read_delivery_file(lca_buf.getvalue(), "lca_test.xlsx", format_key="RROI")
    raise AssertionError("an LCA file under the RROI format must fail (Media_Cost is missing)")
except ValueError as exc:
    assert "Media_Cost" in str(exc)
print("TEST H OK: the LCA format reads Media Cost / Video Views by their spaced names, "
      "and is rejected under RROI (Media_Cost missing)")

# ---------------------------------------------------------------------------
# TEST I: a brand spanning several categories keeps them all — the category is
# never silently standardized down to the first one
# ---------------------------------------------------------------------------
multi_cat_rows = [
    row("2026-06-01", brand="Dove", category="Hair Care", placement="PL1_UNE",
        impressions=None, media_cost=5.0),
    row("2026-06-01", brand="Dove", category="Skin Care", placement="PL2_UNE",
        impressions=100, media_cost=None),
]
df_multi, *_ = read_delivery_file(make_upload_bytes(multi_cat_rows), "multi.xlsx")
cats = sorted(df_multi.loc[df_multi["Brand"] == "Dove", "Category"].unique())
assert cats == ["Hair Care", "Skin Care"], cats
# each row keeps its own category through detection — nothing is rewritten
caveats_multi = classify(build_groups(df_multi, False, JUNE, JUNE))
assert sorted(caveats_multi["Category"].unique()) == ["Hair Care", "Skin Care"], \
    "detection must preserve each row's own category"
# and the joined label is what reaches the tab name, rather than just the first category
tab_multi = build_sheet_name("INC1", " & ".join(cats), "Dove")
assert "HairCare" in tab_multi and "SkinCare" in tab_multi, tab_multi
assert len(tab_multi) <= 31
print("TEST I OK: a multi-category brand keeps every category; the tab name names them all")

# ---------------------------------------------------------------------------
# TEST J: the Month column carries the year — written as a real date, formatted
# "mmm-yy", and two same-month/different-year lines stay distinct
# ---------------------------------------------------------------------------
from datetime import datetime as _dt  # noqa: E402

two_year_rows = [
    # Jun 2025 and Jun 2026, same placement: a bare "Jun" would collapse them
    row("2025-06-01", placement="PLY_UNE", impressions=None, media_cost=5.0),
    row("2026-06-01", placement="PLY_UNE", impressions=None, media_cost=9.0),
]
df_years, *_ = read_delivery_file(make_upload_bytes(two_year_rows), "years.xlsx")
grouped_years = build_groups(df_years, False, pd.Period("2025-06", freq="M"), JUNE)
caveats_years = classify(grouped_years)
assert len(caveats_years) == 2, "the two years must stay two separate caveat lines"
assert set(caveats_years["Month"]) == {pd.Timestamp("2025-06-01"), pd.Timestamp("2026-06-01")}, \
    set(caveats_years["Month"])

# and the written cell is a real date with the "mmm-yy" number format, not text
years_bytes, _ = write_brand_file_bytes(
    caveats_years, "BrandA", "Hair Care", template_bytes, "INC1234567"
)
ws_years = openpyxl.load_workbook(BytesIO(years_bytes)).worksheets[0]
written_months = [ws_years.cell(row=r, column=9) for r in (8, 9)]
assert all(isinstance(c.value, _dt) for c in written_months), \
    [(c.value, type(c.value).__name__) for c in written_months]
assert all(c.number_format == MONTH_CELL_FORMAT for c in written_months), \
    [c.number_format for c in written_months]
assert {c.value.year for c in written_months} == {2025, 2026}
print("TEST J OK: Month is written as a real date formatted 'mmm-yy', keeping the year "
      "so Jun'25 and Jun'26 remain distinct lines")

print("ALL DATA CAVEATS TESTS PASSED")
