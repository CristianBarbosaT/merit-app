"""Merit Deliver — builds the client-facing deliverable, reconciles it against the
source, and flags visual duplicates.

Ported from the standalone script `Merit_Deliver.py` (file-picker + local `output/`
folder, run by hand) into an in-app, upload/download flow. All file I/O happens in
memory (BytesIO); nothing is written to disk. Business rules (DELIVERABLE_SCHEMA,
reconciliation logic, duplicate classification) are unchanged from the original script.
"""
import re
import zipfile
from datetime import datetime
from io import BytesIO

import openpyxl
import pandas as pd
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment

# ============================================================================
# Fixed business rules (same as the standalone script — edit here, not in render())
# ============================================================================
DELIVERABLE_SCHEMA = [
    ("Channel",                   "Channel"),
    ("Date",                      "Date"),
    ("Brand",                     "Brand"),
    ("Product_Line",              "Product_Line"),
    ("Campaign",                  "Campaign"),
    ("Prisma_Campaign_Secondary", "Prisma_Campaign_Secondary"),
    ("Category",                  "Category"),
    ("Raw_Partner",               "Raw_Partner"),
    ("Audience",                  "Audience"),
    ("Package_Placement_Name",    "Package_Placement_Name"),
    ("Daypart",                   "Daypart"),
    ("Breakout",                  "Breakout"),
    ("Retailer",                  "Retailer"),
    ("Impressions",               "Impressions"),
    ("Clicks",                    "Clicks"),
    ("Media_Cost",                "Media_Cost"),
    ("Video_Views",               "Video_Views"),
    ("GRPs",                      "GRPs"),
]
SOURCE_COLS = [src for _, src in DELIVERABLE_SCHEMA]
BACKUP_EXTRA_COLUMNS = ["Creative Name", "Network_Name"]
RECON_METRICS = [
    ("Impressions", "Impressions", "Impressions"),
    ("Cost",        "Media_Cost",  "Media_Cost"),
    ("GRPs",        "GRPs",        "GRPs"),
]
RECON_DIMS = ["Channel", "Product_Line"]
DATE_COL = "Date"
DATE_OUTPUT_FORMAT = "m/d/yyyy"
BENIGN_DIFFERENTIATORS = ["Creative Name", "CCD JTBD", "Network_Name"]
ALLOW_MISSING_AS_BLANK = False
DUP_EXCLUDE_ZERO_IMPR_AND_SPEND = True
DUP_IMPR_COL  = "Impressions"
DUP_SPEND_COL = "Media_Cost"
EPS = 0.005

FILL_HEADER = PatternFill("solid", fgColor="83D0C6")  # monte carlo
FILL_GREEN  = PatternFill("solid", fgColor="ADC865")  # wild willow
FILL_YELLOW = PatternFill("solid", fgColor="FAC172")  # rojah
FILL_RED    = PatternFill("solid", fgColor="E25B45")  # flame pea

MONTH_NAMES = {1: "January", 2: "February", 3: "March", 4: "April", 5: "May", 6: "June",
               7: "July", 8: "August", 9: "September", 10: "October", 11: "November", 12: "December"}

STATE_DEFAULTS = {
    "deliver_results": None,   # dict produced by the "Generate" step
}


def init_state():
    for key, value in STATE_DEFAULTS.items():
        if key not in st.session_state:
            st.session_state[key] = value


def reset_all():
    for key, value in STATE_DEFAULTS.items():
        st.session_state[key] = value


# ----------------------------------------------------------------------------
# Small utilities (unchanged logic from the standalone script)
# ----------------------------------------------------------------------------
def _sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/:*?"|]', "_", name).strip()


def _period_label(p):
    if p is pd.NaT or pd.isna(p):
        return "Unknown"
    return f"{MONTH_NAMES[p.month]} {p.year}"


def _parse_dates(s: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(s):
        return s
    out = pd.to_datetime(s, format="%m/%d/%Y", errors="coerce")
    bad = out.isna() & s.notna()
    if bad.any():
        out = out.fillna(pd.to_datetime(s, errors="coerce"))
    return out


def _validate_benign(df) -> list:
    """Non-fatal check: warns (instead of printing) if a BENIGN_DIFFERENTIATORS
    name doesn't match a real column — it would be silently ignored otherwise."""
    unknown = [c for c in BENIGN_DIFFERENTIATORS if c not in df.columns]
    if unknown:
        return ["These benign differentiators aren't real columns and will be ignored: "
                + ", ".join(unknown) + ". Check the exact names in your header."]
    return []


# ----------------------------------------------------------------------------
# Core business logic (unchanged from the standalone script)
# ----------------------------------------------------------------------------
def build_deliverable(df: pd.DataFrame) -> pd.DataFrame:
    """18 columns, in order, renamed. keep_all: every row preserved, nothing dropped."""
    missing = [src for src in SOURCE_COLS if src not in df.columns]
    if missing and not ALLOW_MISSING_AS_BLANK:
        hints = []
        for m in missing:
            norm = m.replace("_", " ").strip().casefold()
            cand = [c for c in df.columns if c.replace("_", " ").strip().casefold() == norm]
            hints.append(f"'{m}'" + (f" -> did you mean '{cand[0]}'?" if cand else " -> not found"))
        raise ValueError(
            "These columns are in DELIVERABLE_SCHEMA but missing from the file:\n" + "\n".join(hints)
        )

    out = pd.DataFrame()
    for output_name, source_name in DELIVERABLE_SCHEMA:
        out[output_name] = df[source_name].values if source_name in df.columns else pd.NA
    if DATE_COL in out.columns:
        out[DATE_COL] = pd.to_datetime(out[DATE_COL], errors="coerce").dt.normalize()
    return out


def build_backup_deliverable(df, deliverable_out):
    backup = deliverable_out.copy()
    warnings = []
    for col in BACKUP_EXTRA_COLUMNS:
        if col in df.columns:
            backup[col] = df[col].values
        else:
            backup[col] = pd.NA
            warnings.append(f"Backup column '{col}' not found -> blank.")
    return backup, warnings


def build_reconciliation(df: pd.DataFrame, deliverable: pd.DataFrame):
    o = df[RECON_DIMS].copy()
    for label, src, _ in RECON_METRICS:
        o[label] = pd.to_numeric(df[src], errors="coerce")
    so = o.groupby(RECON_DIMS, dropna=False).agg(
        Rows=(RECON_METRICS[0][0], "size"),
        **{f"{lbl}_o": (lbl, "sum") for lbl, _, _ in RECON_METRICS})

    d = deliverable[RECON_DIMS].copy()
    for label, _, out in RECON_METRICS:
        d[label] = pd.to_numeric(deliverable[out], errors="coerce")
    sd = d.groupby(RECON_DIMS, dropna=False).agg(
        **{f"{lbl}_d": (lbl, "sum") for lbl, _, _ in RECON_METRICS})

    rec = so.join(sd, how="outer").fillna(0)
    rec = rec.reset_index()

    diffs = {}
    leak = pd.Series(False, index=rec.index)
    for lbl, _, _ in RECON_METRICS:
        diff = (rec[f"{lbl}_o"] - rec[f"{lbl}_d"]).abs()
        diffs[lbl] = diff
        leak = leak | (diff > EPS)

    def _status(idx):
        if not leak[idx]:
            return "OK - Exact match"
        bad = [lbl for lbl, _, _ in RECON_METRICS if diffs[lbl][idx] > EPS]
        return "CHECK: " + ", ".join(bad)

    rec["Transfer"] = [_status(i) for i in rec.index]

    tidy = pd.DataFrame({
        "Channel":            rec[RECON_DIMS[0]],
        "Product Line":       rec[RECON_DIMS[1]],
        "Rows":               rec["Rows"].astype(int),
        "Total Impressions":  rec[f"{RECON_METRICS[0][0]}_d"],
        "Total Cost":         rec[f"{RECON_METRICS[1][0]}_d"],
        "Total GRPs":         rec[f"{RECON_METRICS[2][0]}_d"],
        "Transfer":           rec["Transfer"],
    }).sort_values(["Channel", "Product Line"]).reset_index(drop=True)

    integ = []
    for label, src, _ in RECON_METRICS:
        raw = df[src]
        coerced = pd.to_numeric(raw, errors="coerce")
        bad = coerced.isna() & raw.notna() & (raw.astype(str).fillna("").str.strip() != "")
        integ.append({"Metric": label, "Non-numeric cells": int(bad.sum()),
                      "Total": float(coerced.sum())})
    integrity = pd.DataFrame(integ)

    totals = {
        "rows_orig": int(so["Rows"].sum()),
        "rows_deliv": int(rec["Rows"].sum()),
        "all_ok": not leak.any(),
        "bad_cells": int(integrity["Non-numeric cells"].sum()),
    }
    return tidy, integrity, totals


def classify_duplicates(df: pd.DataFrame):
    n_excluded = 0
    scan = df
    if DUP_EXCLUDE_ZERO_IMPR_AND_SPEND:
        impr = pd.to_numeric(df[DUP_IMPR_COL], errors="coerce").fillna(0)
        cost = pd.to_numeric(df[DUP_SPEND_COL], errors="coerce").fillna(0)
        empty = (impr == 0) & (cost == 0)
        n_excluded = int(empty.sum())
        scan = df[~empty]

    dropped = [c for c in df.columns if c not in SOURCE_COLS]
    records = []
    for keys, g in scan.groupby(SOURCE_COLS, dropna=False, sort=False):
        if len(g) < 2:
            continue
        keys = keys if isinstance(keys, tuple) else (keys,)
        varying   = [c for c in dropped if g[c].nunique(dropna=False) > 1]
        nonbenign = [c for c in varying if c not in BENIGN_DIFFERENTIATORS]

        if not varying:
            sev, verdict = "TRUE DUP", "Identical across ALL source columns"
        elif nonbenign:
            sev, verdict = "REVIEW", "Differ by non-benign field(s): " + ", ".join(nonbenign)
        else:
            sev, verdict = "EXPECTED", "Differ only by: " + ", ".join(varying)

        rows = ", ".join(str(i + 2) for i in g.index[:25]) + (", ..." if len(g) > 25 else "")
        rec = {"Severity": sev, "Verdict": verdict, "Identical_rows": len(g),
               "Differentiated_by": ", ".join(varying) or "(nothing)",
               "Source_Excel_Rows": rows}
        for sc in ["Channel", "Brand", "Product_Line", "Package_Placement_Name"]:
            if sc in SOURCE_COLS:
                rec[sc] = keys[SOURCE_COLS.index(sc)]
        records.append(rec)

    cols = ["Severity", "Verdict", "Identical_rows", "Differentiated_by",
            "Source_Excel_Rows", "Channel", "Brand", "Product_Line", "Package_Placement_Name"]
    out = pd.DataFrame(records, columns=cols)
    if len(out):
        order = {"REVIEW": 0, "TRUE DUP": 1, "EXPECTED": 2}
        out = out.sort_values("Severity", key=lambda s: s.map(order)).reset_index(drop=True)
    return out, n_excluded


def scan_formulas(file_bytes: bytes, sheet_index: int = 0) -> pd.DataFrame:
    """Scans every cell of the given sheet — every column, not just the metric
    columns — and returns one row per cell holding a LIVE formula instead of a
    plain value (e.g. "=SUM(A1:A2)"). pd.read_excel already returns the cached
    calculated value for formula cells, so this needs its own pass over the raw
    workbook with data_only=False to see the formula text itself."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False, read_only=True)
    try:
        ws = wb.worksheets[sheet_index]
        header = None
        records = []
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            if r_idx == 1:
                header = [c.value for c in row]
                continue
            for c_idx, cell in enumerate(row):
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    col_name = (header[c_idx] if header and c_idx < len(header) and header[c_idx]
                                else f"Column {c_idx + 1}")
                    records.append({"Column": col_name, "Excel_Row": r_idx, "Formula": value})
    finally:
        wb.close()
    return pd.DataFrame(records, columns=["Column", "Excel_Row", "Formula"])


def build_output_filename(df):
    brands = sorted({str(b).strip() for b in df["Brand"].dropna() if str(b).strip()})
    brand_part = brands[0] if len(brands) == 1 else ("NoBrand" if not brands else "Multi-Brand")
    dates = _parse_dates(df["Date"]).dropna()
    if len(dates):
        s, e = dates.min().to_period("M"), dates.max().to_period("M")
        span = _period_label(s) if s == e else f"{_period_label(s)} - {_period_label(e)}"
    else:
        span = "Unknown period"
    return _sanitize_filename(f"{brand_part} {span}.xlsx")


# ----------------------------------------------------------------------------
# Writing outputs — to BytesIO instead of disk paths
# ----------------------------------------------------------------------------
def _write_delivery_sheet_bytes(frame, sheet_name="Deliverable") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]
        for c in ws[1]:
            c.font = Font(bold=True); c.fill = FILL_HEADER
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        if DATE_COL in frame.columns:
            idx = list(frame.columns).index(DATE_COL) + 1
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=idx).number_format = DATE_OUTPUT_FORMAT
    return output.getvalue()


def _bruno_deliver_comment(totals, n_orig, dups):
    n_true = int((dups["Severity"] == "TRUE DUP").sum()) if len(dups) else 0
    n_rev  = int((dups["Severity"] == "REVIEW").sum())   if len(dups) else 0
    parts = [f"{n_orig:,} rows delivered."]
    parts.append("All metrics reconciled - nothing leaked." if totals["all_ok"]
                 else "Metric leak detected - check the reconciliation table.")
    if totals["bad_cells"] > 0:
        parts.append(f"{totals['bad_cells']} non-numeric metric cell(s) found.")
    flag_dups = n_true + n_rev
    parts.append(f"{flag_dups} duplicate group(s) worth a look before you send."
                 if flag_dups > 0 else "No duplicates to review.")
    return "  ".join(parts)


def write_qa_file_bytes(recon, integrity, dups, totals, input_filename, n_orig, n_dup_excluded) -> bytes:
    n_exp  = int((dups["Severity"] == "EXPECTED").sum())  if len(dups) else 0
    n_rev  = int((dups["Severity"] == "REVIEW").sum())    if len(dups) else 0
    n_true = int((dups["Severity"] == "TRUE DUP").sum())  if len(dups) else 0

    bruno_line = _bruno_deliver_comment(totals, n_orig, dups)

    headline = pd.DataFrame({
        "Field": ["Input file", "Run timestamp", "Mode",
                  "Original rows", "Delivered rows", "Rows preserved",
                  "Rows excluded from dup scan (0 impr & 0 spend)",
                  "Metric reconciliation", "Non-numeric metric cells",
                  "Duplicates - Expected", "  - Review", "  - True duplicate"],
        "Value": [input_filename, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "keep_all",
                  f"{n_orig:,}", f"{totals['rows_deliv']:,}",
                  "YES" if n_orig == totals["rows_deliv"] else "NO - CHECK",
                  f"{n_dup_excluded:,}",
                  "OK - all totals match" if totals["all_ok"] else "LEAK - CHECK",
                  f"{totals['bad_cells']:,}",
                  f"{n_exp:,}", f"{n_rev:,}", f"{n_true:,}"],
    })

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        headline.to_excel(writer, sheet_name="Summary", index=False, startrow=2)
        r1 = 2 + len(headline) + 3
        integrity.to_excel(writer, sheet_name="Summary", index=False, startrow=r1)
        r2 = r1 + len(integrity) + 3
        recon.to_excel(writer, sheet_name="Summary", index=False, startrow=r2)
        dups.to_excel(writer, sheet_name="Expected_Duplicates", index=False)

        ws = writer.sheets["Summary"]
        ws.merge_cells("A1:G1")
        bc = ws.cell(row=1, column=1, value=bruno_line)
        bc.font = Font(italic=True, size=10, color="444444")
        bc.fill = PatternFill("solid", fgColor="FAFAFA")
        bc.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
        ws.row_dimensions[1].height = 30

        hl_header = 3
        for row in (hl_header, r1 + 1, r2 + 1):
            for c in ws[row]:
                c.font = Font(bold=True); c.fill = FILL_HEADER

        def _color_field(label, fill):
            for r in range(hl_header + 1, hl_header + 1 + len(headline)):
                if ws.cell(row=r, column=1).value == label:
                    ws.cell(row=r, column=2).fill = fill
                    return
        _color_field("Metric reconciliation", FILL_GREEN if totals["all_ok"] else FILL_RED)
        _color_field("Non-numeric metric cells", FILL_GREEN if totals["bad_cells"] == 0 else FILL_YELLOW)

        for r in range(r1 + 2, r1 + 2 + len(integrity)):
            ws.cell(row=r, column=3).number_format = "#,##0.00"

        recon_first = r2 + 2
        for r in range(recon_first, recon_first + len(recon)):
            ws.cell(row=r, column=3).number_format = "#,##0"
            ws.cell(row=r, column=4).number_format = "#,##0"
            ws.cell(row=r, column=5).number_format = "#,##0.00"
            ws.cell(row=r, column=6).number_format = "#,##0"
            tcell = ws.cell(row=r, column=7)
            if tcell.value and str(tcell.value).startswith("OK"):
                tcell.fill = FILL_GREEN
            elif tcell.value:
                tcell.fill = FILL_RED

        for col, w in {"A": 24, "B": 20, "C": 20, "D": 20, "E": 16, "F": 12, "G": 22}.items():
            ws.column_dimensions[col].width = w

        wp = writer.sheets["Expected_Duplicates"]
        for c in wp[1]:
            c.font = Font(bold=True); c.fill = FILL_HEADER
        wp.freeze_panes = "A2"; wp.auto_filter.ref = wp.dimensions
        color = {"EXPECTED": FILL_GREEN, "REVIEW": FILL_YELLOW, "TRUE DUP": FILL_RED}
        for r in range(2, wp.max_row + 1):
            fill = color.get(str(wp.cell(row=r, column=1).value))
            if fill:
                for cc in range(1, wp.max_column + 1):
                    wp.cell(row=r, column=cc).fill = fill
        wp.column_dimensions["B"].width = 46
        wp.column_dimensions["E"].width = 40

    return output.getvalue()


# ----------------------------------------------------------------------------
# In-app styling for the QA tables (same color logic as the Excel version)
# ----------------------------------------------------------------------------
def _style_recon(recon: pd.DataFrame):
    def highlight(row):
        color = FILL_GREEN.fgColor.rgb if str(row["Transfer"]).startswith("OK") else FILL_RED.fgColor.rgb
        return [f"background-color: #{str(color)[-6:]}; color: #1a1a1a"] * len(row)
    return recon.style.apply(highlight, axis=1).format({
        "Total Impressions": "{:,.0f}", "Total Cost": "{:,.2f}", "Total GRPs": "{:,.0f}",
    })


def _style_integrity(integrity: pd.DataFrame):
    def highlight(row):
        color = "" if row["Non-numeric cells"] == 0 else f"background-color: #{str(FILL_YELLOW.fgColor.rgb)[-6:]}; color: #1a1a1a"
        return [color] * len(row)
    return integrity.style.apply(highlight, axis=1).format({"Total": "{:,.2f}"})


def _style_dups(dups: pd.DataFrame):
    color_map = {"EXPECTED": FILL_GREEN, "REVIEW": FILL_YELLOW, "TRUE DUP": FILL_RED}
    def highlight(row):
        fill = color_map.get(row["Severity"])
        color = f"background-color: #{str(fill.fgColor.rgb)[-6:]}; color: #1a1a1a" if fill else ""
        return [color] * len(row)
    return dups.style.apply(highlight, axis=1)


def _style_formula_summary(summary: pd.DataFrame):
    def highlight(row):
        color = f"background-color: #{str(FILL_YELLOW.fgColor.rgb)[-6:]}; color: #1a1a1a" if row["Formula cells"] > 0 else ""
        return [color] * len(row)
    return summary.style.apply(highlight, axis=1)


# ----------------------------------------------------------------------------
# UI
# ----------------------------------------------------------------------------
def render():
    if st.button("← Back to menu", key="deliver_back"):
        st.session_state.active_tool = None
        st.rerun()

    init_state()

    st.title("Merit Deliver")
    st.caption(
        "Builds the client-facing deliverable from your raw file, reconciles totals "
        "against the source (Channel x Product Line), and flags visual duplicates "
        "before you send anything out."
    )

    uploaded = st.file_uploader(
        "Upload the raw Excel file", type=["xlsx", "xls"], key="deliver_uploader",
    )

    if uploaded and st.button("Generate deliverable", type="primary", use_container_width=True):
        try:
            df = pd.read_excel(uploaded, sheet_name=0, engine="openpyxl").reset_index(drop=True)
            benign_warnings = _validate_benign(df)
            deliverable = build_deliverable(df)
            recon, integrity, totals = build_reconciliation(df, deliverable)
            dups, n_excluded = classify_duplicates(df)
            backup, backup_warnings = build_backup_deliverable(df, deliverable)

            try:
                formulas = scan_formulas(uploaded.getvalue(), sheet_index=0)
            except Exception as exc:
                formulas = pd.DataFrame(columns=["Column", "Excel_Row", "Formula"])
                benign_warnings = benign_warnings + [f"Formula scan failed: {exc}"]

            base = build_output_filename(df)
            deliv_bytes  = _write_delivery_sheet_bytes(deliverable, "Deliverable")
            backup_bytes = _write_delivery_sheet_bytes(backup, "Backup")
            qa_bytes     = write_qa_file_bytes(recon, integrity, dups, totals, uploaded.name,
                                                len(df), n_excluded)

            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr(f"RROI Delivery {base}", deliv_bytes)
                zf.writestr(f"Backup RROI Delivery {base}", backup_bytes)
                zf.writestr(f"RROI Delivery QA {base}", qa_bytes)

            st.session_state.deliver_results = {
                "fatal_error": None,
                "base": base,
                "n_orig": len(df),
                "totals": totals,
                "n_excluded": n_excluded,
                "n_exp":  int((dups["Severity"] == "EXPECTED").sum()) if len(dups) else 0,
                "n_rev":  int((dups["Severity"] == "REVIEW").sum())   if len(dups) else 0,
                "n_true": int((dups["Severity"] == "TRUE DUP").sum()) if len(dups) else 0,
                "warnings": benign_warnings + backup_warnings,
                "recon": recon,
                "integrity": integrity,
                "dups": dups,
                "formulas": formulas,
                "zip_bytes": zip_buffer.getvalue(),
            }
        except Exception as exc:
            st.session_state.deliver_results = {"fatal_error": str(exc)}
        st.rerun()

    results = st.session_state.deliver_results
    if not results:
        return

    st.divider()
    st.subheader("Results")

    if results.get("fatal_error"):
        st.error(results["fatal_error"])
        return

    for w in results["warnings"]:
        st.warning(w)

    totals = results["totals"]
    if totals["all_ok"]:
        st.success(f"{results['n_orig']:,} rows processed. All metrics reconciled — nothing leaked.")
    else:
        st.error(f"{results['n_orig']:,} rows processed. Metric leak detected — check the QA file.")

    if totals["bad_cells"] > 0:
        st.warning(f"{totals['bad_cells']} non-numeric metric cell(s) — delivered as blank.")

    if results["n_rev"] or results["n_true"]:
        st.warning(f"{results['n_rev'] + results['n_true']} duplicate group(s) worth a look before you send.")
    else:
        st.info("No visual duplicates in the deliverable.")

    if results["n_excluded"]:
        st.caption(f"{results['n_excluded']:,} empty rows excluded from the dup scan (0 impr & 0 spend).")

    st.divider()
    st.subheader("Reconciliation — Channel x Product Line")
    st.dataframe(_style_recon(results["recon"]), use_container_width=True, hide_index=True)

    st.subheader("Integrity scan — non-numeric metric cells")
    st.dataframe(_style_integrity(results["integrity"]), use_container_width=True, hide_index=True)

    st.subheader("Formula scan — cells with a live formula instead of a plain value")
    formulas = results["formulas"]
    if len(formulas):
        summary = (formulas.groupby("Column").size()
                   .reset_index(name="Formula cells")
                   .sort_values("Formula cells", ascending=False))
        st.warning(
            f"{len(formulas):,} cell(s) across {len(summary)} column(s) hold a formula, not a "
            "plain value — if the file is re-saved without recalculating, or opened somewhere "
            "the linked source is missing, those cells can silently go stale or blank."
        )
        st.dataframe(_style_formula_summary(summary), use_container_width=True, hide_index=True)
        with st.expander(f"See all {len(formulas):,} flagged cells"):
            st.dataframe(formulas, use_container_width=True, hide_index=True)
    else:
        st.caption("No live formulas found — every cell is a plain value.")

    st.subheader("Duplicate groups")
    if len(results["dups"]):
        st.dataframe(_style_dups(results["dups"]), use_container_width=True, hide_index=True)
    else:
        st.caption("No visual duplicates in the deliverable.")

    st.download_button(
        "Download deliverable, backup and QA (.zip)",
        data=results["zip_bytes"],
        file_name=f"MeritDeliver_{_sanitize_filename(results['base'])}.zip",
        mime="application/zip",
        type="primary",
    )
